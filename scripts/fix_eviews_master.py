import pandas as pd
import numpy as np
import json
import re
from openpyxl import load_workbook

# ─── Load corrected data ───────────────────────────────────────────────
wb_path = 'Food_VAT_EViews_Master_2008_2025_v3.xlsx'
price_kj = pd.read_excel(wb_path, sheet_name='Price_per_100kJ')
lookup = pd.read_excel(wb_path, sheet_name='Product_Lookup')

with open('classification_data.json') as f:
    classif = json.load(f)

price_kj['Date'] = pd.to_datetime(price_kj['Date'])
zr_codes = set(int(k) for k in classif['zero_rated'].keys())

# Extract COICOP codes from Price_per_100kJ column headers
code_to_col = {}
for c in price_kj.columns[1:]:
    m = re.search(r'\((\d+)\)', c)
    if m:
        code_to_col[int(m.group(1))] = c

# Build lookup dict (now with CORRECTED categories)
lookup_dict = {}
for _, row in lookup.iterrows():
    lookup_dict[row['COICOP_Code']] = {
        'name': row['Product_Name'],
        'category': row['Food_Category'],
        'zero_rated': row['Zero_Rated'] == 'Yes',
        'kj': row['kJ_per_100g'],
        'weight': row['CPI_Weight'] if pd.notna(row['CPI_Weight']) else 0
    }

# ─── Recalculate category averages ─────────────────────────────────────────
categories = ['Starchy foods', 'Fruit & vegetables', 'Dairy & eggs',
              'Meat, fish & poultry', 'Fats & oils', 'Sugar & sweets',
              'Processed foods', 'Beverages']

cat_series = {cat: [] for cat in categories}
zr_all = []
all_food = []

for code, col in code_to_col.items():
    if code not in lookup_dict:
        continue
    info = lookup_dict[code]
    cat = info['category']
    
    series = price_kj.set_index('Date')[col]
    
    if cat in cat_series:
        cat_series[cat].append(series)
    
    all_food.append(series)
    
    if code in zr_codes:
        zr_all.append(series)

# Compute averages
cat_avg = {}
for cat, series_list in cat_series.items():
    if series_list:
        combined = pd.concat(series_list, axis=1)
        cat_avg[cat] = combined.mean(axis=1)
        print(f"  {cat}: {len(series_list)} products, "
              f"R{cat_avg[cat].dropna().iloc[0]:.4f} → R{cat_avg[cat].dropna().iloc[-1]:.4f}")

zr_avg = pd.concat(zr_all, axis=1).mean(axis=1)
all_avg = pd.concat(all_food, axis=1).mean(axis=1)

print(f"\n  Zero-rated: {len(zr_all)} products")
print(f"  All food: {len(all_food)} products")

# ─── Update EViews_Master sheet ──────────────────────────────────────────────
wb = load_workbook(wb_path)
ws = wb['EViews_Master']

# Find column indices
headers = {}
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    if val:
        headers[val] = col

print(f"\nEViews_Master headers: {headers}")

col_mapping = {
    'Starchy_foods': 'Starchy foods',
    'Fruit_and_vegetables': 'Fruit & vegetables',
    'Dairy_and_eggs': 'Dairy & eggs',
    'Meat_fish_and_poultry': 'Meat, fish & poultry',
    'Fats_and_oils': 'Fats & oils',
    'Sugar_and_sweets': 'Sugar & sweets',
    'Processed_foods': 'Processed foods',
}

# Update each category column
dates = price_kj['Date'].tolist()
for excel_col_name, cat_name in col_mapping.items():
    if excel_col_name in headers and cat_name in cat_avg:
        col_idx = headers[excel_col_name]
        avg = cat_avg[cat_name]
        for row_idx, date in enumerate(dates, start=2):
            if date in avg.index and pd.notna(avg[date]):
                ws.cell(row=row_idx, column=col_idx).value = round(avg[date], 6)
            # Leave existing value if no data (shouldn't happen)
        print(f"  Updated {excel_col_name}")

# Update Zero-rated column
if 'Zero-rated' in headers:
    col_idx = headers['Zero-rated']
    for row_idx, date in enumerate(dates, start=2):
        if date in zr_avg.index and pd.notna(zr_avg[date]):
            ws.cell(row=row_idx, column=col_idx).value = round(zr_avg[date], 6)
    print("  Updated Zero-rated")

# Update All_food column
if 'All_food' in headers:
    col_idx = headers['All_food']
    for row_idx, date in enumerate(dates, start=2):
        if date in all_avg.index and pd.notna(all_avg[date]):
            ws.cell(row=row_idx, column=col_idx).value = round(all_avg[date], 6)
    print("  Updated All_food")

# ─── Add Beverages column if not present ─────────────────────────────────
if 'Beverages' not in headers and 'Beverages' in cat_avg:
    # Insert after Processed_foods
    proc_col = headers.get('Processed_foods', 10)
    # Actually, let's not insert columns as it could shift everything
    # Just note this for the user
    print("\n  NOTE: Beverages category not in EViews_Master (wasn't there before)")

# ─── Save ──────────────────────────────────────────────────────────────────────
wb.save(wb_path)
print(f"\nSaved updated {wb_path}")

# ─── Verify ─────────────────────────────────────────────────────────────────────
eviews = pd.read_excel(wb_path, sheet_name='EViews_Master')
print("\n=== VERIFICATION ===")
for col_name, cat_name in col_mapping.items():
    if col_name in eviews.columns:
        first = eviews[col_name].dropna().iloc[0]
        last = eviews[col_name].dropna().iloc[-1]
        print(f"  {col_name}: R{first:.4f} → R{last:.4f}")
print(f"  Zero-rated: R{eviews['Zero-rated'].dropna().iloc[0]:.4f} → R{eviews['Zero-rated'].dropna().iloc[-1]:.4f}")
print(f"  All_food: R{eviews['All_food'].dropna().iloc[0]:.4f} → R{eviews['All_food'].dropna().iloc[-1]:.4f}")
