import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

# ─── COICOP-based category mapping (Stats SA official classification) ────
COICOP_TO_CAT = {
    # 011.1 Bread and cereals → Starchy foods
    1111: 'Starchy foods', 1112: 'Starchy foods', 1113: 'Starchy foods',
    1114: 'Starchy foods', 1115: 'Starchy foods',
    # 011.2 Meat → Meat, fish & poultry
    1122: 'Meat, fish & poultry', 1123: 'Meat, fish & poultry',
    1124: 'Meat, fish & poultry', 1125: 'Meat, fish & poultry',
    # 011.3 Fish and seafood → Meat, fish & poultry
    1131: 'Meat, fish & poultry', 1132: 'Meat, fish & poultry',
    1133: 'Meat, fish & poultry', 1134: 'Meat, fish & poultry',
    # 011.4 Milk, cheese, eggs → Dairy & eggs
    1141: 'Dairy & eggs', 1142: 'Dairy & eggs', 1143: 'Dairy & eggs',
    1145: 'Dairy & eggs', 1146: 'Dairy & eggs', 1147: 'Dairy & eggs',
    1148: 'Dairy & eggs',
    # 011.5 Oils and fats → Fats & oils
    1151: 'Fats & oils', 1152: 'Fats & oils', 1153: 'Fats & oils',
    # 011.6 Fruit → Fruit & vegetables
    1161: 'Fruit & vegetables', 1163: 'Fruit & vegetables',
    1165: 'Fruit & vegetables', 1169: 'Fruit & vegetables',
    # 011.7 Vegetables → Fruit & vegetables
    1171: 'Fruit & vegetables', 1172: 'Fruit & vegetables',
    1174: 'Fruit & vegetables', 1175: 'Fruit & vegetables',
    1176: 'Fruit & vegetables', 1178: 'Fruit & vegetables',
    1179: 'Fruit & vegetables',
    # 011.8 Sugar, jam, honey, chocolate, confectionery → Sugar & sweets
    1181: 'Sugar & sweets', 1183: 'Sugar & sweets', 1184: 'Sugar & sweets',
    1185: 'Sugar & sweets', 1186: 'Sugar & sweets', 1189: 'Sugar & sweets',
    # 011.9 Food products n.e.c. → Processed foods
    1191: 'Processed foods', 1192: 'Processed foods', 1193: 'Processed foods',
    1194: 'Processed foods', 1199: 'Processed foods',
    # 012 Non-alcoholic beverages → Beverages
    1210: 'Beverages', 1220: 'Beverages', 1230: 'Beverages',
    1250: 'Beverages', 1260: 'Beverages', 1290: 'Beverages',
}

# ─── Load workbook (preserve all formatting) ─────────────────────────────
wb = load_workbook('Food_VAT_EViews_Master_2008_2025_v2.xlsx')
ws = wb['Product_Lookup']

# Find the Food_Category column
header_row = 1
cat_col = None
code_col = None
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=header_row, column=col).value
    if val == 'Food_Category':
        cat_col = col
    elif val == 'COICOP_Code':
        code_col = col

print(f"COICOP_Code column: {code_col}")
print(f"Food_Category column: {cat_col}")

changes = []
for row in range(2, ws.max_row + 1):
    code_val = ws.cell(row=row, column=code_col).value
    if code_val is None:
        continue
    code = int(code_val)
    subclass = int(str(code)[:4])
    correct_cat = COICOP_TO_CAT.get(subclass)
    current_cat = ws.cell(row=row, column=cat_col).value
    
    if correct_cat and correct_cat != current_cat:
        product_name = ws.cell(row=row, column=2).value
        changes.append((row, product_name, current_cat, correct_cat))
        ws.cell(row=row, column=cat_col).value = correct_cat

print(f"\nTotal corrections made: {len(changes)}")
for row_num, name, old, new in sorted(changes, key=lambda x: x[2]):
    print(f"  Row {row_num}: {name}: '{old}' → '{new}'")

# ─── Also update the Overview sheet if it has category counts ─────────────
if 'Overview' in wb.sheetnames:
    print("\nOverview sheet exists - will need manual review for category counts")

# ─── Save corrected workbook ──────────────────────────────────────────
output = 'Food_VAT_EViews_Master_2008_2025_v3.xlsx'
wb.save(output)
print(f"\nSaved corrected file: {output}")

# ─── Verify the corrections ────────────────────────────────────────────
df = pd.read_excel(output, sheet_name='Product_Lookup')
print("\n=== VERIFIED CATEGORY COUNTS ===")
for cat in sorted(df['Food_Category'].unique()):
    n = (df['Food_Category'] == cat).sum()
    zr = ((df['Food_Category'] == cat) & (df['Zero_Rated'] == 'Yes')).sum()
    print(f"  {cat}: {n} items ({zr} zero-rated)")

# Verify no "Condiments & other" remains
if 'Condiments & other' in df['Food_Category'].values:
    print("\nWARNING: 'Condiments & other' still present!")
else:
    print("\n'Condiments & other' eliminated (all items properly classified by COICOP)")
