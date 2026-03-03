#!/usr/bin/env python3
"""
Build the final EViews-ready Excel workbook and publication-quality graph.
Uses intermediate data from build_master_data.py, fixes inf values by
excluding products with 0 kJ (salt, tea, water).
"""

import pandas as pd
import numpy as np
import json
import warnings
warnings.filterwarnings('ignore')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

###############################################################################
# RE-LOAD AND FIX DATA (exclude 0 kJ products)
###############################################################################

df_cpi = pd.read_excel('EXCEL-CPI-COICOP-2018-8digit-202512.xlsx', sheet_name=0)
food_cpi = df_cpi[df_cpi['Division'].astype(str).str.strip() == '1'].copy()
month_cols = [c for c in df_cpi.columns if str(c).startswith('M')]

df_prices = pd.read_excel('CPI_Average-Prices_All-urban-202512.xlsx', sheet_name=0)
price_month_cols = [c for c in df_prices.columns if str(c).startswith('M')]

with open('kj_per_100g_database.json') as f:
    kj_db = json.load(f)

import re

def parse_unit_grams(unit_str):
    if pd.isna(unit_str):
        return None
    s = str(unit_str).strip().lower().replace(',', '.')
    m = re.match(r'([\d.]+)\s*(kilogram|kg|gram|g|litre|liter|l|ml|millilitre)', s)
    if m:
        val = float(m.group(1))
        unit = m.group(2)
        if unit in ['kilogram', 'kg']: return val * 1000
        elif unit in ['gram', 'g']: return val
        elif unit in ['litre', 'liter', 'l']: return val * 1000
        elif unit in ['ml', 'millilitre']: return val
    return None

# Build price per 100g
df_prices['grams'] = df_prices['H08'].apply(parse_unit_grams)
food_prices = df_prices[df_prices['H03'].notna() & df_prices['grams'].notna()].copy()

price_per_100g = {}
for code, grp in food_prices.groupby('H03'):
    code = int(code)
    smallest = grp.loc[grp['grams'].idxmin()]
    grams = smallest['grams']
    monthly_prices = {}
    for mc in price_month_cols:
        val = smallest[mc]
        if pd.notna(val) and isinstance(val, (int, float)):
            monthly_prices[mc] = (val / grams) * 100
    if monthly_prices:
        price_per_100g[code] = {'prices': monthly_prices}

# Back-calculate 2008-2016 prices
full_prices_per_100g = {}
for _, row in food_cpi.iterrows():
    code = int(row['Eight digit code']) if pd.notna(row['Eight digit code']) else None
    if code is None: continue
    name = str(row['Product name']).strip()
    index_values = {}
    for mc in month_cols:
        val = row[mc]
        if pd.notna(val) and isinstance(val, (int, float)) and val > 0:
            index_values[mc] = val
    if not index_values: continue
    if code in price_per_100g:
        actual_prices = price_per_100g[code]['prices']
        ref_month = None
        for mc in sorted(price_month_cols):
            if mc in actual_prices and mc in index_values:
                ref_month = mc
                break
        if ref_month:
            ref_price = actual_prices[ref_month]
            ref_index = index_values[ref_month]
            computed_prices = {}
            for mc in month_cols:
                if mc in actual_prices:
                    computed_prices[mc] = actual_prices[mc]
                elif mc in index_values:
                    computed_prices[mc] = ref_price * (index_values[mc] / ref_index)
            full_prices_per_100g[code] = {'name': name, 'prices': computed_prices}

# Classification
def classify_product(name, subclass):
    if subclass in [1111, 1112, 1113, 1114, 1115]: return 'Starchy foods'
    elif subclass in [1161, 1163, 1165, 1169]: return 'Fruit & vegetables'
    elif subclass in [1141, 1143, 1145, 1146, 1147, 1148]: return 'Dairy & eggs'
    elif subclass in [1122, 1123, 1124, 1125, 1131, 1132, 1133, 1134]: return 'Meat, fish & poultry'
    elif subclass in [1151, 1152, 1153]: return 'Fats & oils'
    elif subclass in [1171, 1172, 1174, 1175, 1176, 1178, 1179]: return 'Sugar & sweets'
    elif subclass in [1181, 1183, 1184, 1185, 1186, 1189]: return 'Condiments & other'
    elif subclass in [1191, 1192, 1193, 1194, 1199]: return 'Processed foods'
    elif subclass in [1210, 1220, 1230, 1250, 1260, 1290]: return 'Beverages'
    return 'Other'

def is_zero_rated(name, subclass, code):
    nl = name.lower()
    if 'brown bread' in nl: return True
    if any(x in nl for x in ['maize meal', 'samp', 'mealie rice', 'sorghum meal']): return True
    if any(x in nl for x in ['cake flour', 'bread flour']): return True
    if 'rice' in nl and subclass == 1111: return True
    if any(x in nl for x in ['dried beans', 'lentils', 'dried peas']): return True
    if any(x in nl for x in ['pilchard', 'sardine']): return True
    if any(x in nl for x in ['fresh milk', 'full cream milk', 'low fat milk', 'fat free milk']): return True
    if 'milk powder' in nl or 'dairy powder' in nl: return True
    if subclass == 1148 or nl.strip() == 'eggs': return True
    if any(x in nl for x in ['vegetable oil', 'sunflower oil', 'cooking oil']): return True
    if subclass in [1161, 1163, 1165]: return True
    if any(x in nl for x in ['tinned vegetable', 'frozen vegetable']): return True
    return False

products = []
for _, row in food_cpi.iterrows():
    code = int(row['Eight digit code']) if pd.notna(row['Eight digit code']) else None
    if code is None: continue
    name = str(row['Product name']).strip()
    subclass = int(row['Subclass']) if pd.notna(row['Subclass']) else None
    weight = row['Weight'] if pd.notna(row['Weight']) else None
    category = classify_product(name, subclass)
    zero_rated = is_zero_rated(name, subclass, code)
    kj_val = None
    nl = name.lower()
    if name in kj_db: kj_val = kj_db[name]
    else:
        for db_name, db_kj in kj_db.items():
            if db_name.lower() == nl:
                kj_val = db_kj; break
        if kj_val is None:
            for db_name, db_kj in kj_db.items():
                if db_name.lower() in nl or nl in db_name.lower():
                    kj_val = db_kj; break
    has_prices = code in full_prices_per_100g
    products.append({
        'code': code, 'name': name, 'subclass': subclass, 'cpi_weight': weight,
        'category': category, 'zero_rated': zero_rated, 'kj_per_100g': kj_val,
        'has_full_prices': has_prices
    })

df_products = pd.DataFrame(products)

# Compute price per 100kJ (EXCLUDE kJ=0 products!)
eligible = df_products[
    (df_products['kj_per_100g'].notna()) & 
    (df_products['kj_per_100g'] > 0) &  # <-- FIX: exclude 0 kJ
    (df_products['has_full_prices'])
].copy()
print(f"Eligible products (kJ>0 and has prices): {len(eligible)}")

price_per_100kj_data = {}
for _, prod in eligible.iterrows():
    code = prod['code']
    kj = prod['kj_per_100g']
    prices = full_prices_per_100g[code]['prices']
    price_per_100kj_data[code] = {mc: (p / kj) * 100 for mc, p in prices.items()}

# Category averages
categories = ['Starchy foods', 'Fruit & vegetables', 'Dairy & eggs', 
              'Meat, fish & poultry', 'Fats & oils', 'Sugar & sweets', 'Processed foods']
all_cats = categories + ['Zero-rated', 'All food']
category_avg = {cat: {} for cat in all_cats}

for mc in month_cols:
    for cat in categories:
        cat_prods = eligible[eligible['category'] == cat]
        vals, wts = [], []
        for _, prod in cat_prods.iterrows():
            code = prod['code']
            if code in price_per_100kj_data and mc in price_per_100kj_data[code]:
                v = price_per_100kj_data[code][mc]
                if np.isfinite(v):
                    vals.append(v)
                    wts.append(prod['cpi_weight'] if pd.notna(prod['cpi_weight']) else 1)
        if vals: category_avg[cat][mc] = np.average(vals, weights=wts)
    
    # Zero-rated
    zr_prods = eligible[eligible['zero_rated']]
    vals, wts = [], []
    for _, prod in zr_prods.iterrows():
        code = prod['code']
        if code in price_per_100kj_data and mc in price_per_100kj_data[code]:
            v = price_per_100kj_data[code][mc]
            if np.isfinite(v):
                vals.append(v)
                wts.append(prod['cpi_weight'] if pd.notna(prod['cpi_weight']) else 1)
    if vals: category_avg['Zero-rated'][mc] = np.average(vals, weights=wts)
    
    # All food
    vals, wts = [], []
    for _, prod in eligible.iterrows():
        code = prod['code']
        if code in price_per_100kj_data and mc in price_per_100kj_data[code]:
            v = price_per_100kj_data[code][mc]
            if np.isfinite(v):
                vals.append(v)
                wts.append(prod['cpi_weight'] if pd.notna(prod['cpi_weight']) else 1)
    if vals: category_avg['All food'][mc] = np.average(vals, weights=wts)

print("Category averages (first and last month):")
for cat in all_cats:
    vals = category_avg[cat]
    if vals:
        first = min(vals.keys())
        last = max(vals.keys())
        print(f"  {cat}: R{vals[first]:.4f} ({first}) -> R{vals[last]:.4f} ({last})")

###############################################################################
# BUILD EVIEWS-READY TIME SERIES
###############################################################################

def mc_to_date(mc):
    return pd.Timestamp(year=int(mc[1:5]), month=int(mc[5:7]), day=1)

# Sheet 1: Category time series
rows = []
for mc in month_cols:
    dt = mc_to_date(mc)
    row = {'Date': dt, 'Year': dt.year, 'Month': dt.month}
    for cat in all_cats:
        col = cat.replace(' ', '_').replace(',', '').replace('&', 'and')
        row[col] = category_avg[cat].get(mc, np.nan)
    rows.append(row)
df_cat_ts = pd.DataFrame(rows)

# FAO
df_fao_raw = pd.read_excel('food_price_indices_data_feb-2.xlsx', 
                            sheet_name='Indices_Monthly_Nominal', header=None, skiprows=4)
df_fao = df_fao_raw.iloc[:,:7].copy()
df_fao.columns = ['Date','FAO_Food_Price_Index','FAO_Meat','FAO_Dairy','FAO_Cereals','FAO_Oils','FAO_Sugar']
df_fao['Date'] = pd.to_datetime(df_fao['Date'], errors='coerce')
df_fao = df_fao.dropna(subset=['Date'])
df_fao_08 = df_fao[(df_fao['Date'] >= '2008-01-01') & (df_fao['Date'] <= '2025-12-31')].copy()
df_fao_08 = df_fao_08.reset_index(drop=True)

# PMBEJD
df_pmb1 = pd.read_excel('pmbejd_affordability_2025-2.xlsx', sheet_name=0)
df_pmb2 = pd.read_excel('pmbejd_affordability_extended.xlsx', sheet_name=0)
df_pmb = pd.concat([df_pmb1, df_pmb2]).drop_duplicates(subset=['date']).sort_values('date').reset_index(drop=True)

# Merge FAO into category time series
df_merged = df_cat_ts.copy()
df_merged = df_merged.merge(df_fao_08, on='Date', how='left')

# Merge PMBEJD
pmb_cols = ['date','food_basket_7p','basket_zero_rated_cost','basket_vatable_cost',
            'basket_vat_value','basket_vat_share','headline_inflation','cpi_food_yoy',
            'nmw_monthly','food_poverty_line','food_shortfall_share']
pmb_available = [c for c in pmb_cols if c in df_pmb.columns]
df_pmb_sel = df_pmb[pmb_available].copy()
df_pmb_sel = df_pmb_sel.rename(columns={'date': 'Date'})
df_merged = df_merged.merge(df_pmb_sel, on='Date', how='left')

print(f"\nMerged master: {df_merged.shape}")
print(f"Columns: {list(df_merged.columns)}")

###############################################################################
# BUILD EXCEL WORKBOOK
###############################################################################
print("\n" + "=" * 60)
print("BUILDING EXCEL WORKBOOK")
print("=" * 60)

wb = Workbook()

# ─── STYLES ───
header_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2B5797')
subheader_fill = PatternFill('solid', fgColor='4472C4')
data_font = Font(name='Arial', size=9)
title_font = Font(name='Arial', bold=True, size=14, color='2B5797')
subtitle_font = Font(name='Arial', size=10, color='666666')
thin_border = Border(
    bottom=Side(style='thin', color='D9D9D9')
)

def write_header(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    return cell

def write_data(ws, row, col, value, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = data_font
    cell.border = thin_border
    if fmt: cell.number_format = fmt
    return cell

# ═══════════════════════════════════════════════════════════════
# SHEET 1: OVERVIEW
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'Overview'
ws1.sheet_properties.tabColor = '2B5797'

ws1.column_dimensions['A'].width = 3
ws1.column_dimensions['B'].width = 45
ws1.column_dimensions['C'].width = 60

r = 2
ws1.cell(row=r, column=2, value='VAT Zero-Rating & Food Affordability').font = title_font
ws1.merge_cells('B2:C2')
r = 3
ws1.cell(row=r, column=2, value='Master Data Workbook for EViews Analysis (2008-2025)').font = subtitle_font
ws1.merge_cells('B3:C3')

r = 5
ws1.cell(row=r, column=2, value='Sheet Guide').font = Font(name='Arial', bold=True, size=12)
r = 7
sheets_info = [
    ('EViews_Master', 'Monthly time series: price per 100kJ by category, FAO indices, PMBEJD data (2008-2025)'),
    ('Product_Lookup', 'Product classification: COICOP codes, food category, zero-rated status, kJ/100g'),
    ('CPI_Indices', 'Raw CPI price indices (base Dec 2024=100) for 134 food products'),
    ('Price_per_100g', 'Rand prices per 100g (actual 2017-2025, back-calculated 2008-2016)'),
    ('Price_per_100kJ', 'Rand prices per 100kJ for each product'),
    ('FAO_Indices', 'FAO Food Price Indices (nominal, base 2014-2016=100)'),
    ('PMBEJD', 'PMBEJD Affordability Basket data'),
]
for sheet_name, desc in sheets_info:
    from openpyxl.worksheet.hyperlink import Hyperlink
    cell = ws1.cell(row=r, column=2, value=sheet_name)
    cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"'{sheet_name}'!A1")
    cell.font = Font(name='Arial', color='0000FF', underline='single', size=10)
    ws1.cell(row=r, column=3, value=desc).font = data_font
    r += 1

r += 2
ws1.cell(row=r, column=2, value='Data Sources').font = Font(name='Arial', bold=True, size=12)
r += 2
sources = [
    'Stats SA CPI COICOP 8-digit indices (P0141, 2008-2025)',
    'Stats SA CPI Average Prices All Urban (P0141, 2017-2025)', 
    'SAFOODS Food Composition Database (kJ per 100g)',
    'FAO Food Price Index (nominal monthly, 1990-2026)',
    'PMBEJD Household Affordability Index (2025-2026)',
    'Stats SA Income & Expenditure Survey 2022/23'
]
for src in sources:
    ws1.cell(row=r, column=2, value=f'• {src}').font = data_font
    r += 1

r += 2
ws1.cell(row=r, column=2, value='Methodology Notes').font = Font(name='Arial', bold=True, size=12)
r += 2
notes = [
    'CPI indices (base Dec 2024=100) combined with actual Rand prices (2017-2025) to back-calculate 2008-2016 prices.',
    'Price per 100g = (Price per pack) / (grams per pack) × 100. Smallest pack size used for each product.',
    'Price per 100kJ = (Price per 100g) / (kJ per 100g) × 100. Products with 0 kJ (salt, tea, water) excluded.',
    'Category averages are CPI-weight-weighted means of constituent products.',
    'Zero-rated classification based on SA VAT Act Schedule 2 Part B (19 basic foodstuffs).',
    f'Total products: {len(df_products)}, with kJ values: {eligible.shape[0]}, zero-rated: {df_products["zero_rated"].sum()}'
]
for note in notes:
    cell = ws1.cell(row=r, column=2, value=note)
    cell.font = data_font
    cell.alignment = Alignment(wrap_text=True)
    ws1.merge_cells(f'B{r}:C{r}')
    ws1.row_dimensions[r].height = 30
    r += 1

# ═══════════════════════════════════════════════════════════════
# SHEET 2: EVIEWS MASTER (main time series)
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('EViews_Master')
ws2.sheet_properties.tabColor = '00B050'

# Write headers
headers2 = list(df_merged.columns)
for c, h in enumerate(headers2, 1):
    write_header(ws2, 1, c, h)
    ws2.column_dimensions[get_column_letter(c)].width = max(14, len(str(h)) + 2)

# Write data
for r, (_, row) in enumerate(df_merged.iterrows(), 2):
    for c, h in enumerate(headers2, 1):
        val = row[h]
        if pd.isna(val):
            write_data(ws2, r, c, None)
        elif h == 'Date':
            write_data(ws2, r, c, val, 'YYYY-MM-DD')
        elif h in ['Year', 'Month']:
            write_data(ws2, r, c, int(val), '#,##0')
        else:
            write_data(ws2, r, c, float(val) if np.isfinite(val) else None, '0.0000')

# Freeze top row
ws2.freeze_panes = 'A2'

# Add table
last_row2 = len(df_merged) + 1
last_col2 = len(headers2)
tab2 = Table(displayName='EViewsMaster', ref=f'A1:{get_column_letter(last_col2)}{last_row2}')
tab2.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True)
ws2.add_table(tab2)

# ═══════════════════════════════════════════════════════════════
# SHEET 3: PRODUCT LOOKUP
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Product_Lookup')
ws3.sheet_properties.tabColor = 'FFC000'

lookup_cols = ['code', 'name', 'subclass', 'category', 'zero_rated', 'kj_per_100g', 'cpi_weight', 'has_full_prices']
lookup_headers = ['COICOP_Code', 'Product_Name', 'Subclass', 'Food_Category', 'Zero_Rated', 'kJ_per_100g', 'CPI_Weight', 'Has_Prices']

for c, h in enumerate(lookup_headers, 1):
    write_header(ws3, 1, c, h)

widths = [14, 35, 10, 22, 12, 14, 12, 12]
for c, w in enumerate(widths, 1):
    ws3.column_dimensions[get_column_letter(c)].width = w

df_prod_sorted = df_products.sort_values(['category', 'name']).reset_index(drop=True)
for r, (_, row) in enumerate(df_prod_sorted.iterrows(), 2):
    for c, col in enumerate(lookup_cols, 1):
        val = row[col]
        if pd.isna(val):
            write_data(ws3, r, c, None)
        elif col == 'zero_rated':
            write_data(ws3, r, c, 'Yes' if val else 'No')
        elif col in ['code', 'subclass']:
            write_data(ws3, r, c, int(val))
        elif col == 'kj_per_100g':
            write_data(ws3, r, c, float(val) if val and np.isfinite(val) else None, '#,##0')
        elif col == 'cpi_weight':
            write_data(ws3, r, c, float(val) if pd.notna(val) else None, '0.00000')
        elif col == 'has_full_prices':
            write_data(ws3, r, c, 'Yes' if val else 'No')
        else:
            write_data(ws3, r, c, str(val))

ws3.freeze_panes = 'A2'
tab3 = Table(displayName='ProductLookup', ref=f'A1:{get_column_letter(len(lookup_headers))}{len(df_products)+1}')
tab3.tableStyleInfo = TableStyleInfo(name='TableStyleMedium6', showFirstColumn=False, showLastColumn=False, showRowStripes=True)
ws3.add_table(tab3)

# Highlight zero-rated rows
zr_fill = PatternFill('solid', fgColor='E2EFDA')
for r in range(2, len(df_prod_sorted) + 2):
    if ws3.cell(row=r, column=5).value == 'Yes':
        for c in range(1, len(lookup_headers) + 1):
            ws3.cell(row=r, column=c).fill = zr_fill

# ═══════════════════════════════════════════════════════════════
# SHEET 4: CPI INDICES (raw)
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('CPI_Indices')
ws4.sheet_properties.tabColor = '4472C4'

# Product codes as columns, months as rows
cpi_products = food_cpi[['Eight digit code', 'Product name']].dropna().reset_index(drop=True)
write_header(ws4, 1, 1, 'Date')
ws4.column_dimensions['A'].width = 12
for c, (_, prod) in enumerate(cpi_products.iterrows(), 2):
    code = int(prod['Eight digit code'])
    name = prod['Product name']
    write_header(ws4, 1, c, f"{name}\n({code})")
    ws4.column_dimensions[get_column_letter(c)].width = 12

for r, mc in enumerate(month_cols, 2):
    dt = mc_to_date(mc)
    write_data(ws4, r, 1, dt, 'YYYY-MM-DD')
    for c, (_, prod) in enumerate(cpi_products.iterrows(), 2):
        code = int(prod['Eight digit code'])
        row_data = food_cpi[food_cpi['Eight digit code'] == code]
        if not row_data.empty:
            val = row_data.iloc[0][mc]
            write_data(ws4, r, c, float(val) if pd.notna(val) else None, '0.00')

ws4.freeze_panes = 'B2'

# ═══════════════════════════════════════════════════════════════
# SHEET 5: PRICE PER 100g
# ═══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet('Price_per_100g')
ws5.sheet_properties.tabColor = 'ED7D31'

price_codes = sorted(full_prices_per_100g.keys())
write_header(ws5, 1, 1, 'Date')
ws5.column_dimensions['A'].width = 12
for c, code in enumerate(price_codes, 2):
    name = full_prices_per_100g[code]['name']
    write_header(ws5, 1, c, f"{name}\n({code})")
    ws5.column_dimensions[get_column_letter(c)].width = 12

for r, mc in enumerate(month_cols, 2):
    dt = mc_to_date(mc)
    write_data(ws5, r, 1, dt, 'YYYY-MM-DD')
    for c, code in enumerate(price_codes, 2):
        val = full_prices_per_100g[code]['prices'].get(mc)
        write_data(ws5, r, c, float(val) if val and np.isfinite(val) else None, '0.00')

ws5.freeze_panes = 'B2'

# ═══════════════════════════════════════════════════════════════
# SHEET 6: PRICE PER 100kJ
# ═══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet('Price_per_100kJ')
ws6.sheet_properties.tabColor = 'FF0000'

pkj_codes = sorted(price_per_100kj_data.keys())
write_header(ws6, 1, 1, 'Date')
ws6.column_dimensions['A'].width = 12
for c, code in enumerate(pkj_codes, 2):
    name = full_prices_per_100g[code]['name'] if code in full_prices_per_100g else str(code)
    write_data(ws6, 1, c, f"{name}\n({code})")
    ws6.cell(row=1, column=c).font = header_font
    ws6.cell(row=1, column=c).fill = header_fill
    ws6.cell(row=1, column=c).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws6.column_dimensions[get_column_letter(c)].width = 12

for r, mc in enumerate(month_cols, 2):
    dt = mc_to_date(mc)
    write_data(ws6, r, 1, dt, 'YYYY-MM-DD')
    for c, code in enumerate(pkj_codes, 2):
        val = price_per_100kj_data[code].get(mc)
        write_data(ws6, r, c, float(val) if val and np.isfinite(val) else None, '0.0000')

ws6.freeze_panes = 'B2'

# ═══════════════════════════════════════════════════════════════
# SHEET 7: FAO INDICES
# ═══════════════════════════════════════════════════════════════
ws7 = wb.create_sheet('FAO_Indices')
ws7.sheet_properties.tabColor = '7030A0'

fao_headers = list(df_fao_08.columns)
for c, h in enumerate(fao_headers, 1):
    write_header(ws7, 1, c, h)
    ws7.column_dimensions[get_column_letter(c)].width = 18

for r, (_, row) in enumerate(df_fao_08.iterrows(), 2):
    for c, h in enumerate(fao_headers, 1):
        val = row[h]
        if h == 'Date':
            write_data(ws7, r, c, val, 'YYYY-MM-DD')
        else:
            write_data(ws7, r, c, float(val) if pd.notna(val) else None, '0.00')

ws7.freeze_panes = 'A2'

# ═══════════════════════════════════════════════════════════════
# SHEET 8: PMBEJD
# ═══════════════════════════════════════════════════════════════
ws8 = wb.create_sheet('PMBEJD')
ws8.sheet_properties.tabColor = '00B0F0'

pmb_headers = list(df_pmb.columns)
for c, h in enumerate(pmb_headers, 1):
    write_header(ws8, 1, c, h)
    ws8.column_dimensions[get_column_letter(c)].width = max(14, len(str(h)) + 2)

for r, (_, row) in enumerate(df_pmb.iterrows(), 2):
    for c, h in enumerate(pmb_headers, 1):
        val = row[h]
        if pd.isna(val):
            write_data(ws8, r, c, None)
        elif h == 'date':
            write_data(ws8, r, c, val, 'YYYY-MM-DD')
        elif h in ['year', 'month']:
            write_data(ws8, r, c, int(val))
        else:
            write_data(ws8, r, c, float(val), '0.00')

ws8.freeze_panes = 'A2'

# ═══════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════
output_path = 'Food_VAT_EViews_Master_2008_2025.xlsx'
wb.save(output_path)
print(f"\nWorkbook saved: {output_path}")
print(f"Sheets: {wb.sheetnames}")

# Save category averages for graphing
with open('_category_avg_fixed.json', 'w') as f:
    json.dump({cat: {k: v for k, v in vals.items()} for cat, vals in category_avg.items()}, f)

print("DONE!")
