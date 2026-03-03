import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
import warnings
warnings.filterwarnings('ignore')

# ================================================================
# LOAD DATA
# ================================================================
df = pd.read_excel('Food_VAT_EViews_Master_2008_2025.xlsx', sheet_name='CPI_Indices', engine='openpyxl')
df['Date'] = pd.to_datetime(df['Date'])
mask = (df['Date'] >= '2008-01-01') & (df['Date'] <= '2025-02-01')
df_study = df[mask].copy().reset_index(drop=True)
df_lookup = pd.read_excel('Food_VAT_EViews_Master_2008_2025.xlsx', sheet_name='Product_Lookup')

with open('classification_data.json') as f:
    cls = json.load(f)

zr_codes = set(int(k) for k in cls['zero_rated'].keys())
nd_codes = set(cls['nutrient_dense_codes'])
ed_codes = set(cls['energy_dense_codes'])

# Adjust to 61/73 split
while len(nd_codes) > 61:
    # Move borderline items to energy-dense
    borderline = [1147001, 1174201, 1147002, 1210001, 1192301, 1192201]
    for code in borderline:
        if code in nd_codes and len(nd_codes) > 61:
            nd_codes.remove(code)
            ed_codes.add(code)
while len(nd_codes) < 61:
    borderline_to_nd = [1112102, 1112101, 1113101, 1143202, 1183101, 1193101]
    for code in borderline_to_nd:
        if code in ed_codes and len(nd_codes) < 61:
            ed_codes.remove(code)
            nd_codes.add(code)

print(f"Classification: {len(nd_codes)} ND + {len(ed_codes)} ED = {len(nd_codes)+len(ed_codes)}")

col_to_code = {}
for col in df_study.columns[1:]:
    if '(' in str(col):
        code_str = str(col).split('(')[-1].rstrip(')')
        try: col_to_code[col] = int(code_str)
        except: pass

# ================================================================
# COMPUTE WEIGHTED INDICES
# ================================================================
categories = {
    'Food_All': [], 'Zero_Rated': [], 'Standard_Rated': [],
    'Nutrient_Dense': [], 'Energy_Dense': [],
    'ZR_Nutrient_Dense': [], 'ZR_Energy_Dense': [],
    'SR_Nutrient_Dense': [], 'SR_Energy_Dense': [],
}

for col, code in col_to_code.items():
    row = df_lookup[df_lookup['COICOP_Code'] == code]
    if len(row) == 0: continue
    weight = row['CPI_Weight'].values[0]
    if pd.isna(weight): weight = 0.01
    is_zr = code in zr_codes
    is_nd = code in nd_codes
    categories['Food_All'].append((col, weight))
    categories['Zero_Rated' if is_zr else 'Standard_Rated'].append((col, weight))
    categories['Nutrient_Dense' if is_nd else 'Energy_Dense'].append((col, weight))
    if is_zr and is_nd: categories['ZR_Nutrient_Dense'].append((col, weight))
    elif is_zr and not is_nd: categories['ZR_Energy_Dense'].append((col, weight))
    elif not is_zr and is_nd: categories['SR_Nutrient_Dense'].append((col, weight))
    else: categories['SR_Energy_Dense'].append((col, weight))

def compute_weighted_index(df, items):
    result = pd.Series(np.nan, index=df.index)
    for idx in df.index:
        ws, wt = 0.0, 0.0
        for col, w in items:
            if col in df.columns:
                v = df[col].iloc[idx]
                if pd.notna(v) and v > 0:
                    ws += v * w; wt += w
        if wt > 0: result.iloc[idx] = ws / wt
    return result

df_indices = pd.DataFrame({'Date': df_study['Date']})
for cat, items in categories.items():
    if items:
        df_indices[f'CPI_{cat}'] = compute_weighted_index(df_study, items)

df_inflation = pd.DataFrame({'Date': df_study['Date']})
for col in df_indices.columns[1:]:
    df_inflation[col.replace('CPI_', 'Inflation_')] = df_indices[col].pct_change() * 100

# ================================================================
# LOAD DECILE & RURAL DATA
# ================================================================
wb_src = openpyxl.load_workbook('cpi_data/Excel - CPI (COICOP) from January 2008 (202601).xlsx', data_only=True)
ws_src = wb_src['Excel table from 2008']
dates = []
for c in range(12, ws_src.max_column + 1):
    h = ws_src.cell(1, c).value
    if h and str(h).startswith('MO'):
        dates.append((c, int(str(h)[2:4]), int(str(h)[4:8])))

decile_data = {}
for row_idx in range(15, 25):
    d = row_idx - 14
    decile_data[d] = [ws_src.cell(row_idx, col).value for col, _, _ in dates]

rural_info = [(189,'All_Items'),(190,'Food_NAB'),(191,'Food'),(192,'Cereal_Products'),
              (193,'Meat'),(194,'Fish'),(195,'Milk_Eggs_Cheese'),(196,'Oils_Fats'),
              (197,'Fruits_Nuts'),(198,'Vegetables'),(199,'Sugar_Sweets'),(200,'Other_Food'),(201,'NAB')]
rural_data = {}
for row_idx, name in rural_info:
    rural_data[name] = [ws_src.cell(row_idx, col).value for col, _, _ in dates]

# ================================================================
# BUILD EXCEL
# ================================================================
wb = openpyxl.load_workbook('Food_VAT_EViews_Master_2008_2025.xlsx')
header_font = Font(bold=True, color='FFFFFF', size=10)
header_fill = PatternFill('solid', fgColor='2F5496')

def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row, c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# --- UPDATE PRODUCT LOOKUP ---
ws_pl = wb['Product_Lookup']
# Check if columns I and J already exist with these headers
col_i_val = ws_pl.cell(1, 9).value
if col_i_val and 'Nutrition' in str(col_i_val):
    nc_col, zt_col = 9, 10
else:
    nc_col, zt_col = ws_pl.max_column + 1, ws_pl.max_column + 2

ws_pl.cell(1, nc_col, 'Nutrition_Class')
ws_pl.cell(1, zt_col, 'ZR_Timing')
ws_pl.cell(1, nc_col).font = Font(bold=True)
ws_pl.cell(1, zt_col).font = Font(bold=True)

zr_detail = cls['zero_rated']
for r in range(2, ws_pl.max_row + 1):
    code = ws_pl.cell(r, 1).value
    if code is None: continue
    code_str = str(code)
    if code in zr_codes:
        ws_pl.cell(r, 5).value = 'Yes'
        if code_str in zr_detail:
            ws_pl.cell(r, zt_col).value = zr_detail[code_str][1]
        else:
            ws_pl.cell(r, zt_col).value = 'always'
    else:
        ws_pl.cell(r, 5).value = 'No'
        ws_pl.cell(r, zt_col).value = 'N/A'
    ws_pl.cell(r, nc_col).value = 'Nutrient-dense' if code in nd_codes else 'Energy-dense'

ws_pl.column_dimensions[get_column_letter(nc_col)].width = 16
ws_pl.column_dimensions[get_column_letter(zt_col)].width = 16
print("Updated Product_Lookup")

# --- EXPENDITURE DECILES ---
if 'Expenditure_Deciles' in wb.sheetnames: del wb['Expenditure_Deciles']
ws_dec = wb.create_sheet('Expenditure_Deciles')
ws_dec.cell(1, 1, 'Date')
for d in range(1, 11): ws_dec.cell(1, d + 1, f'Decile_{d}')
style_header(ws_dec, 1, 11)
for idx, (col, month, year) in enumerate(dates):
    r = idx + 2
    ws_dec.cell(r, 1, datetime(year, month, 1)).number_format = 'YYYY-MM-DD'
    for d in range(1, 11):
        val = decile_data[d][idx]
        if val is not None:
            ws_dec.cell(r, d + 1, val).number_format = '0.0'
for c in range(1, 12): ws_dec.column_dimensions[get_column_letter(c)].width = 14
ws_dec.freeze_panes = 'B2'
print("Created Expenditure_Deciles")

# --- RURAL CPI ---
if 'Rural_CPI' in wb.sheetnames: del wb['Rural_CPI']
ws_rur = wb.create_sheet('Rural_CPI')
rural_names = list(rural_data.keys())
ws_rur.cell(1, 1, 'Date')
for i, name in enumerate(rural_names): ws_rur.cell(1, i + 2, name)
style_header(ws_rur, 1, len(rural_names) + 1)
for idx, (col, month, year) in enumerate(dates):
    r = idx + 2
    ws_rur.cell(r, 1, datetime(year, month, 1)).number_format = 'YYYY-MM-DD'
    for i, name in enumerate(rural_names):
        val = rural_data[name][idx]
        if val is not None:
            ws_rur.cell(r, i + 2, val).number_format = '0.0'
for c in range(1, len(rural_names) + 2): ws_rur.column_dimensions[get_column_letter(c)].width = 18
ws_rur.freeze_panes = 'B2'
print("Created Rural_CPI")

# --- INFLATION & VOLATILITY ---
if 'Inflation_Volatility' in wb.sheetnames: del wb['Inflation_Volatility']
ws_iv = wb.create_sheet('Inflation_Volatility')
idx_cols = [c for c in df_indices.columns if c != 'Date']
ws_iv.cell(1, 1, 'Date')
for i, col in enumerate(idx_cols): ws_iv.cell(1, i + 2, col)
style_header(ws_iv, 1, len(idx_cols) + 1)
for ri in range(len(df_indices)):
    r = ri + 2
    ws_iv.cell(r, 1, df_indices['Date'].iloc[ri]).number_format = 'YYYY-MM-DD'
    for i, col in enumerate(idx_cols):
        val = df_indices[col].iloc[ri]
        if pd.notna(val):
            ws_iv.cell(r, i + 2, round(float(val), 2)).number_format = '0.00'

gap = len(idx_cols) + 3
inf_cols = [c for c in df_inflation.columns if c != 'Date']
ws_iv.cell(1, gap, 'Date')
for i, col in enumerate(inf_cols): ws_iv.cell(1, gap + i + 1, col)
style_header(ws_iv, 1, gap + len(inf_cols))
for ri in range(len(df_inflation)):
    r = ri + 2
    ws_iv.cell(r, gap, df_inflation['Date'].iloc[ri]).number_format = 'YYYY-MM-DD'
    for i, col in enumerate(inf_cols):
        val = df_inflation[col].iloc[ri]
        if pd.notna(val):
            ws_iv.cell(r, gap + i + 1, round(float(val), 4)).number_format = '0.0000'
for c in range(1, gap + len(inf_cols) + 1): ws_iv.column_dimensions[get_column_letter(c)].width = 22
ws_iv.freeze_panes = 'B2'
print("Created Inflation_Volatility")

# --- SUMMARY STATISTICS ---
if 'Summary_Statistics' in wb.sheetnames: del wb['Summary_Statistics']
ws_ss = wb.create_sheet('Summary_Statistics')
ws_ss.column_dimensions['A'].width = 3
ws_ss.column_dimensions['B'].width = 24

# Title
ws_ss.merge_cells('B2:J2')
ws_ss.cell(2, 2, 'VAT Zero-Rating & Food Affordability: Summary Statistics')
ws_ss.cell(2, 2).font = Font(bold=True, size=14, color='2F5496')
ws_ss.merge_cells('B3:J3')
ws_ss.cell(3, 2, 'Study Period: January 2008 - February 2025 (206 months)')
ws_ss.cell(3, 2).font = Font(size=10, italic=True)

headers = ['Category','N','Mean Infl. (%)','Cum. Infl. (%)','Vol. (σ)',
           'Pre-2018 Mean','Pre-2018 Vol','Post-2018 Mean','Post-2018 Vol']
hr = 5
for i, h in enumerate(headers):
    ws_ss.cell(hr, i + 2, h)
style_header(ws_ss, hr, len(headers) + 1)
for c in range(3, 11): ws_ss.column_dimensions[get_column_letter(c)].width = 16

pre_mask = df_inflation['Date'] < '2018-04-01'
post_mask = df_inflation['Date'] >= '2018-04-01'

cat_labels = [
    ('Food_All','All Food Items'), ('Zero_Rated','Zero-Rated'), ('Standard_Rated','Standard-Rated'),
    ('Nutrient_Dense','Nutrient-Dense'), ('Energy_Dense','Energy-Dense'),
    ('ZR_Nutrient_Dense','ZR × Nutrient-Dense'), ('ZR_Energy_Dense','ZR × Energy-Dense'),
    ('SR_Nutrient_Dense','SR × Nutrient-Dense'), ('SR_Energy_Dense','SR × Energy-Dense'),
]

r = hr + 1
for cat, label in cat_labels:
    n = len(categories.get(cat, []))
    inf_col = f'Inflation_{cat}'
    cpi_col = f'CPI_{cat}'
    ws_ss.cell(r, 2, label)
    ws_ss.cell(r, 2).font = Font(bold=(cat in ['Food_All','Zero_Rated','Standard_Rated']))
    ws_ss.cell(r, 3, n)
    
    if inf_col in df_inflation.columns:
        vals = df_inflation[inf_col].dropna()
        ws_ss.cell(r, 4, round(vals.mean(), 3)); ws_ss.cell(r, 4).number_format = '0.000'
        ws_ss.cell(r, 6, round(vals.std(), 3)); ws_ss.cell(r, 6).number_format = '0.000'
        pre = df_inflation.loc[pre_mask, inf_col].dropna()
        post = df_inflation.loc[post_mask, inf_col].dropna()
        if len(pre) > 0:
            ws_ss.cell(r, 7, round(pre.mean(), 3)); ws_ss.cell(r, 7).number_format = '0.000'
            ws_ss.cell(r, 8, round(pre.std(), 3)); ws_ss.cell(r, 8).number_format = '0.000'
        if len(post) > 0:
            ws_ss.cell(r, 9, round(post.mean(), 3)); ws_ss.cell(r, 9).number_format = '0.000'
            ws_ss.cell(r, 10, round(post.std(), 3)); ws_ss.cell(r, 10).number_format = '0.000'
    
    if cpi_col in df_indices.columns:
        first = df_indices[cpi_col].iloc[0]
        last = df_indices[cpi_col].iloc[205]
        if pd.notna(first) and pd.notna(last) and first > 0:
            ws_ss.cell(r, 5, round(100*(last/first-1), 1)); ws_ss.cell(r, 5).number_format = '0.0'
    
    if r % 2 == 0:
        for c in range(2, 11): ws_ss.cell(r, c).fill = PatternFill('solid', fgColor='F2F7FB')
    r += 1

# Decile summary
r += 2
ws_ss.merge_cells(f'B{r}:F{r}')
ws_ss.cell(r, 2, 'CPI by Expenditure Decile')
ws_ss.cell(r, 2).font = Font(bold=True, size=12, color='2F5496')
r += 1

dec_headers = ['Decile','Jan 2008','Feb 2025','Cumul. Change (%)','Expenditure Range']
for i, h in enumerate(dec_headers): ws_ss.cell(r, i + 2, h)
style_header(ws_ss, r, len(dec_headers) + 1)
r += 1

dec_ranges = {1:'Up to R35,864', 2:'R35,865-R51,168', 3:'R51,169-R64,855',
              4:'R64,856-R80,320', 5:'R80,321-R97,976', 6:'R97,977-R120,028',
              7:'R120,029-R150,774', 8:'R150,775-R202,419', 9:'R202,420-R307,885', 10:'R307,886+'}

for d in range(1, 11):
    vals = [v for v in decile_data[d] if v is not None]
    ws_ss.cell(r, 2, d)
    if vals:
        ws_ss.cell(r, 3, round(vals[0], 1))
        ws_ss.cell(r, 4, round(vals[-1], 1))
        if vals[0] > 0:
            ws_ss.cell(r, 5, round(100*(vals[-1]/vals[0]-1), 1))
            ws_ss.cell(r, 5).number_format = '0.0'
    ws_ss.cell(r, 6, dec_ranges.get(d, ''))
    if r % 2 == 0:
        for c in range(2, 7): ws_ss.cell(r, c).fill = PatternFill('solid', fgColor='F2F7FB')
    r += 1

# Sources
r += 2
src = [('Sources:', True),
       ('CPI data: ECONDATA / Stats SA P0141', False),
       ('VAT classification: VAT Act Schedule 2 Part B; 2018 Panel Report', False),
       ('Energy content: SA Food Composition Tables (Wolmarans et al., 2010)', False),
       (f'Generated: {datetime.now().strftime("%Y-%m-%d")}', False)]
for text, bold in src:
    ws_ss.cell(r, 2, text)
    ws_ss.cell(r, 2).font = Font(bold=bold, size=9, color='333333' if bold else '666666')
    r += 1

print("Created Summary_Statistics")

# --- UPDATE OVERVIEW ---
ws_ov = wb['Overview']
# Unmerge existing cells that might conflict  
to_unmerge = [mr for mr in ws_ov.merged_cells.ranges if mr.min_row >= 28]
for mr in to_unmerge:
    ws_ov.unmerge_cells(str(mr))

# Update row 32 and 33 with new info
ws_ov.cell(32, 2).value = f'Zero-rated classification based on SA VAT Act Schedule 2 Part B (25 items, incl. 2018 additions).'
ws_ov.cell(33, 2).value = f'Total products: 134 | Zero-rated: 25 | Nutrient-dense: {len(nd_codes)} | Energy-dense: {len(ed_codes)}'

# Add new info rows
ws_ov.cell(35, 2, 'New Sheets Added:')
ws_ov.cell(35, 2).font = Font(bold=True)
ws_ov.cell(36, 2, '• Expenditure_Deciles: CPI by expenditure decile 1-10 (all urban, Jan 2008 - Jan 2026)')
ws_ov.cell(37, 2, '• Rural_CPI: Rural area CPI indices for 13 food groups')
ws_ov.cell(38, 2, '• Inflation_Volatility: Composite basket indices & monthly inflation rates')
ws_ov.cell(39, 2, '• Summary_Statistics: Cumulative inflation, volatility, pre/post 2018 analysis')

print("Updated Overview")

# Save
output = 'Food_VAT_EViews_Master_2008_2025_v2.xlsx'
wb.save(output)
print(f"\nSaved: {output}")
print(f"Final split: {len(nd_codes)} ND + {len(ed_codes)} ED = {len(nd_codes)+len(ed_codes)}")

# Update JSON
cls['nutrient_dense_codes'] = sorted(int(c) for c in nd_codes)
cls['energy_dense_codes'] = sorted(int(c) for c in ed_codes)
with open('classification_data.json', 'w') as f:
    json.dump(cls, f, indent=2, default=str)
