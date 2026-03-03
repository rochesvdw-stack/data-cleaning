import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, date
import json
import calendar

# ============================================================
# LOAD ALL DATA FROM EXISTING v3-3 FILE
# ============================================================
print("Loading existing master file...")
wb_src = openpyxl.load_workbook('Food_VAT_EViews_Master_2008_2025_v3-3.xlsx', data_only=True, read_only=True)

# Products/Classification
ws = wb_src['Product_Lookup']
prod_headers = None
products = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        prod_headers = list(row)
        continue
    products.append(list(row))

# CPI Indices (134 items)
ws = wb_src['CPI_Indices']
cpi_headers = []
cpi_data = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        cpi_headers = [str(h).replace('\n', ' ').strip() if h else '' for h in row]
    else:
        cpi_data.append(list(row))

# Composite indices (Inflation_Volatility)
ws = wb_src['Inflation_Volatility']
iv_headers = []
iv_data = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        iv_headers = list(row)
    else:
        iv_data.append(list(row))

# Expenditure Deciles
ws = wb_src['Expenditure_Deciles']
dec_headers = []
dec_data = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        dec_headers = list(row)
    else:
        dec_data.append(list(row))

# Rural CPI
ws = wb_src['Rural_CPI']
rural_headers = []
rural_data = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        rural_headers = list(row)
    else:
        rural_data.append(list(row))

# FAO Indices
ws = wb_src['FAO_Indices']
fao_headers = []
fao_data = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        fao_headers = list(row)
    else:
        fao_data.append(list(row))

# EViews_Master (for PMBEJD/policy data)
ws = wb_src['EViews_Master']
ev_headers = []
ev_data = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        ev_headers = list(row)
    else:
        ev_data.append(list(row))

# PMBEJD
ws = wb_src['PMBEJD']
pmb_headers = []
pmb_data = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        pmb_headers = list(row)
    else:
        pmb_data.append(list(row))

# Price per 100g
ws = wb_src['Price_per_100g']
p100g_headers = []
p100g_data = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        p100g_headers = [str(h).replace('\n', ' ').strip() if h else '' for h in row]
    else:
        p100g_data.append(list(row))

# Price per 100kJ
ws = wb_src['Price_per_100kJ']
p100kj_headers = []
p100kj_data = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        p100kj_headers = [str(h).replace('\n', ' ').strip() if h else '' for h in row]
    else:
        p100kj_data.append(list(row))

wb_src.close()
print(f"Loaded: {len(products)} products, {len(cpi_data)} CPI months, {len(fao_data)} FAO months")

# ============================================================
# BUILD NMW TIME SERIES (monthly, 2008-2025)
# ============================================================
# NMW was introduced 1 Jan 2019. Before that, use sectoral minimum wages as proxy
# NMW per hour values (official)
nmw_hourly = {
    2019: 20.00, 2020: 20.76, 2021: 21.69, 2022: 23.19,
    2023: 25.42, 2024: 27.58, 2025: 28.79
}
# Pre-NMW: approximate minimum for domestic/farm workers
pre_nmw_hourly = {
    2008: 6.38, 2009: 7.06, 2010: 7.50, 2011: 8.00,
    2012: 8.78, 2013: 9.63, 2014: 10.37, 2015: 11.24,
    2016: 12.42, 2017: 13.37, 2018: 15.00
}
# Monthly NMW = hourly * 8h * 21.67 working days
def nmw_monthly(year):
    if year in nmw_hourly:
        return round(nmw_hourly[year] * 8 * 21.67, 2)
    elif year in pre_nmw_hourly:
        return round(pre_nmw_hourly[year] * 8 * 21.67, 2)
    return None

# ============================================================
# BUILD CSG TIME SERIES (monthly, 2008-2025)
# ============================================================
# CSG amounts from UCT Children Count data
csg_values = {
    (2008, 1): 210, (2008, 4): 210, (2008, 8): 220,
    (2009, 4): 240, (2010, 4): 250, (2011, 4): 260, (2011, 10): 270,
    (2012, 4): 280, (2013, 4): 290, (2013, 10): 300,
    (2014, 4): 310, (2014, 10): 320,
    (2015, 4): 330, (2016, 4): 350, (2016, 10): 360,
    (2017, 4): 380, (2018, 4): 400, (2018, 10): 410,
    (2019, 4): 420, (2019, 10): 430,
    (2020, 4): 440, (2020, 10): 450,
    (2021, 4): 460, (2022, 4): 480,
    (2023, 4): 500, (2023, 10): 510,
    (2024, 4): 530, (2025, 4): 560
}

def get_csg(year, month):
    # Find the latest CSG value at or before this year/month
    best_val = None
    best_key = (0, 0)
    for (y, m), v in csg_values.items():
        if (y, m) <= (year, month) and (y, m) > best_key:
            best_key = (y, m)
            best_val = v
    return best_val

# ============================================================
# BUILD OLD AGE GRANT TIME SERIES
# ============================================================
oag_values = {
    (2008, 4): 940, (2009, 4): 1010, (2010, 4): 1080,
    (2011, 4): 1140, (2012, 4): 1200, (2013, 4): 1260, (2013, 10): 1270,
    (2014, 4): 1350, (2015, 4): 1410, (2016, 4): 1500, (2016, 10): 1510,
    (2017, 4): 1600, (2018, 4): 1690, (2018, 10): 1700,
    (2019, 4): 1780, (2020, 4): 1860,
    (2021, 4): 1890, (2022, 4): 1980, (2022, 10): 1990,
    (2023, 4): 2080, (2023, 10): 2090,
    (2024, 4): 2180, (2024, 10): 2190,
    (2025, 4): 2310
}

def get_oag(year, month):
    best_val = None
    best_key = (0, 0)
    for (y, m), v in oag_values.items():
        if (y, m) <= (year, month) and (y, m) > best_key:
            best_key = (y, m)
            best_val = v
    return best_val

# ============================================================
# FOOD POVERTY LINE (Stats SA, annual)
# ============================================================
food_poverty_line = {
    2008: 210, 2009: 237, 2010: 261, 2011: 305,
    2012: 321, 2013: 335, 2014: 355, 2015: 415,
    2016: 498, 2017: 531, 2018: 547, 2019: 561,
    2020: 585, 2021: 624, 2022: 663, 2023: 760,
    2024: 796, 2025: 796
}

# ============================================================
# SRD GRANT (COVID Social Relief of Distress)
# ============================================================
srd_values = {
    (2020, 5): 350, (2021, 4): 350, (2022, 4): 350,
    (2023, 4): 350, (2024, 4): 370, (2025, 4): 370
}

def get_srd(year, month):
    if year < 2020 or (year == 2020 and month < 5):
        return None
    best_val = None
    best_key = (0, 0)
    for (y, m), v in srd_values.items():
        if (y, m) <= (year, month) and (y, m) > best_key:
            best_key = (y, m)
            best_val = v
    return best_val

# ============================================================
# STRUCTURAL BREAK DUMMY (April 2018 VAT increase)
# ============================================================
def vat_dummy(year, month):
    return 1 if (year > 2018) or (year == 2018 and month >= 4) else 0

# ============================================================
# COVID DUMMY
# ============================================================
def covid_dummy(year, month):
    if (year == 2020 and month >= 3) or (year == 2021 and month <= 3):
        return 1
    return 0

# ============================================================
# CREATE THE PhD EXCEL WORKBOOK
# ============================================================
print("\nBuilding PhD EViews Master Dataset...")
wb = openpyxl.Workbook()

# Styles
header_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
title_font = Font(name='Arial', bold=True, size=14, color='2F5496')
subtitle_font = Font(name='Arial', bold=True, size=11, color='2F5496')
data_font = Font(name='Arial', size=10)
thin_border = Border(
    left=Side(style='thin', color='D9E2F3'),
    right=Side(style='thin', color='D9E2F3'),
    top=Side(style='thin', color='D9E2F3'),
    bottom=Side(style='thin', color='D9E2F3')
)

def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def auto_width(ws, min_w=10, max_w=22):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)

# ============================================================
# SHEET 1: OVERVIEW
# ============================================================
ws_ov = wb.active
ws_ov.title = 'Overview'
ws_ov.sheet_properties.tabColor = '2F5496'

overview_content = [
    [''],
    ['', 'PhD EViews Master Dataset'],
    ['', 'VAT Zero-Rating, Nutrition-Relevant Food Price Inflation,'],
    ['', 'and Healthy Diet Affordability in South Africa: 2008-2025'],
    [''],
    ['', 'Student:', 'Rochelle van der Walt (49218786)'],
    ['', 'Supervisor:', 'Prof. Waldo Krugell'],
    ['', 'Institution:', 'North-West University'],
    ['', 'Period:', 'January 2008 - February 2025 (206 months)'],
    ['', 'Items:', '134 food items (25 zero-rated, 109 standard-rated)'],
    ['', 'Classification:', '61 nutrient-dense, 73 energy-dense'],
    [''],
    ['', 'SHEET INDEX'],
    ['', 'Sheet', 'Description', 'Paper', 'Rows x Cols'],
    ['', 'P1_CPI_134Items', '8-digit COICOP monthly indices for all 134 food items', 'Paper 1', f'{len(cpi_data)} x 134'],
    ['', 'P1_Composite_Indices', 'Weighted composite indices by VAT x Nutrition', 'Paper 1', f'{len(iv_data)} x 9'],
    ['', 'P1_Decile_CPI', 'Expenditure decile CPI indices', 'Paper 1', f'{len(dec_data)} x 10'],
    ['', 'P1_Rural_CPI', 'Rural area CPI by food group', 'Paper 1', f'{len(rural_data)} x 13'],
    ['', 'P2_ARDL_Variables', 'ARDL bounds-testing variables (NRFPI, macro)', 'Paper 2', f'{len(cpi_data)} x 25+'],
    ['', 'P2_Policy_Income', 'NMW, CSG, OAG, SRD, food poverty line', 'Paper 2', f'{len(cpi_data)} x 12'],
    ['', 'P2_FAO_International', 'FAO Food Price Index (international comparison)', 'Paper 2', f'{len(fao_data)} x 6'],
    ['', 'P2_PMBEJD', 'PMBEJD Household Affordability baskets', 'Paper 2', f'{len(pmb_data)} x 20+'],
    ['', 'P2_Price_per_100g', 'Standardised prices per 100g for 134 items', 'Paper 2', f'{len(p100g_data)} x 134'],
    ['', 'P2_Price_per_100kJ', 'Standardised prices per 100kJ for 134 items', 'Paper 2', f'{len(p100kj_data)} x 134'],
    ['', 'P3_IES_Summary', 'IES 2022/23 expenditure summary by quintile', 'Paper 3', '5 x 15'],
    ['', 'Classification', 'Full 134-item classification (VAT, nutrition, COICOP)', 'All', '134 x 15'],
    ['', 'Metadata', 'Variable definitions, sources, EViews commands', 'All', ''],
    [''],
    ['', 'THREE-PAPER STRUCTURE'],
    ['', 'Paper 1:', 'CPI price dynamics of 134 food items (descriptive, structural breaks at April 2018)'],
    ['', 'Paper 2:', 'Nutrition-Relevant Food Price Index (Laspeyres, ARDL bounds testing, Zivot-Andrews)'],
    ['', 'Paper 3:', 'Microsimulation (IES 2022/23 microdata, 19,940 households, counterfactual scenarios)'],
    [''],
    ['', 'EViews IMPORT'],
    ['', 'wfcreate(wf=PhD_VAT) m 2008m1 2025m2'],
    ['', 'read(t=xlsx, s=P1_CPI_134Items, t1=Date) "PhD_EViews_Master_2008_2025.xlsx"'],
    [''],
    ['', f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'],
]

for r, row_data in enumerate(overview_content, 1):
    for c, val in enumerate(row_data):
        cell = ws_ov.cell(row=r, column=c + 1, value=val)
        cell.font = data_font
        cell.border = thin_border

ws_ov.cell(row=2, column=2).font = title_font
ws_ov.cell(row=3, column=2).font = subtitle_font
ws_ov.cell(row=4, column=2).font = subtitle_font
ws_ov.cell(row=13, column=2).font = subtitle_font
ws_ov.cell(row=29, column=2).font = subtitle_font
ws_ov.cell(row=34, column=2).font = subtitle_font

# Merge title cells
ws_ov.merge_cells('B2:D2')
ws_ov.merge_cells('B3:D3')
ws_ov.merge_cells('B4:D4')

ws_ov.column_dimensions['A'].width = 3
ws_ov.column_dimensions['B'].width = 25
ws_ov.column_dimensions['C'].width = 60
ws_ov.column_dimensions['D'].width = 15
ws_ov.column_dimensions['E'].width = 15

# Style index header
for c in range(2, 6):
    cell = ws_ov.cell(row=14, column=c)
    cell.font = header_font
    cell.fill = header_fill

print("  Overview sheet done")

# ============================================================
# SHEET 2: P1_CPI_134Items (Paper 1 main data)
# ============================================================
ws_cpi = wb.create_sheet('P1_CPI_134Items')
ws_cpi.sheet_properties.tabColor = '4472C4'

# Build clean headers: Date + 134 item codes
# Use clean variable names for EViews
cpi_col_names = ['Date']
for h in cpi_headers[1:]:
    # Extract code from parentheses
    code = ''
    if '(' in h and ')' in h:
        code = h.split('(')[-1].replace(')', '').strip()
    name = h.split('(')[0].strip() if '(' in h else h
    clean_name = f"F{code}" if code else name.replace(' ', '_')
    cpi_col_names.append(clean_name)

# Write header
for c, name in enumerate(cpi_col_names, 1):
    ws_cpi.cell(row=1, column=c, value=name)
style_header(ws_cpi, 1, len(cpi_col_names))

# Write data
for r, row_data in enumerate(cpi_data, 2):
    dt = row_data[0]
    if hasattr(dt, 'strftime'):
        ws_cpi.cell(row=r, column=1, value=dt.strftime('%Y-%m'))
    else:
        ws_cpi.cell(row=r, column=1, value=str(dt)[:7])
    for c, val in enumerate(row_data[1:], 2):
        cell = ws_cpi.cell(row=r, column=c)
        if val is not None:
            try:
                cell.value = float(val)
                cell.number_format = '0.0'
            except (ValueError, TypeError):
                cell.value = val

ws_cpi.freeze_panes = 'B2'
print(f"  P1_CPI_134Items: {len(cpi_data)} rows x {len(cpi_col_names)} cols")

# ============================================================
# SHEET 3: P1_Composite_Indices (Paper 1 composite)
# ============================================================
ws_comp = wb.create_sheet('P1_Composite_Indices')
ws_comp.sheet_properties.tabColor = '4472C4'

comp_cols = ['Date', 'IDX_All_Food', 'IDX_Zero_Rated', 'IDX_Standard_Rated',
             'IDX_Nutrient_Dense', 'IDX_Energy_Dense',
             'IDX_ZR_NutrDense', 'IDX_ZR_EnergyDense',
             'IDX_SR_NutrDense', 'IDX_SR_EnergyDense',
             'Year', 'Month', 'Post_VAT_2018', 'COVID_Dummy']

for c, name in enumerate(comp_cols, 1):
    ws_comp.cell(row=1, column=c, value=name)
style_header(ws_comp, 1, len(comp_cols))

for r, row_data in enumerate(iv_data, 2):
    dt = row_data[0]
    if hasattr(dt, 'strftime'):
        date_str = dt.strftime('%Y-%m')
        yr = dt.year
        mn = dt.month
    else:
        date_str = str(dt)[:7]
        yr = int(date_str[:4])
        mn = int(date_str[5:7])
    
    ws_comp.cell(row=r, column=1, value=date_str)
    # CPI composite indices (columns 1-9 from iv_data)
    for c_idx in range(1, 10):
        val = row_data[c_idx] if c_idx < len(row_data) else None
        if val is not None:
            try:
                ws_comp.cell(row=r, column=c_idx + 1, value=float(val))
                ws_comp.cell(row=r, column=c_idx + 1).number_format = '0.00'
            except (ValueError, TypeError):
                pass
    
    ws_comp.cell(row=r, column=11, value=yr)
    ws_comp.cell(row=r, column=12, value=mn)
    ws_comp.cell(row=r, column=13, value=vat_dummy(yr, mn))
    ws_comp.cell(row=r, column=14, value=covid_dummy(yr, mn))

ws_comp.freeze_panes = 'B2'
print(f"  P1_Composite_Indices: {len(iv_data)} rows")

# ============================================================
# SHEET 4: P1_Decile_CPI
# ============================================================
ws_dec = wb.create_sheet('P1_Decile_CPI')
ws_dec.sheet_properties.tabColor = '4472C4'

dec_col_names = ['Date', 'CPSD_01', 'CPSD_02', 'CPSD_03', 'CPSD_04', 'CPSD_05',
                 'CPSD_06', 'CPSD_07', 'CPSD_08', 'CPSD_09', 'CPSD_10']
for c, name in enumerate(dec_col_names, 1):
    ws_dec.cell(row=1, column=c, value=name)
style_header(ws_dec, 1, len(dec_col_names))

for r, row_data in enumerate(dec_data, 2):
    dt = row_data[0]
    if hasattr(dt, 'strftime'):
        ws_dec.cell(row=r, column=1, value=dt.strftime('%Y-%m'))
    else:
        ws_dec.cell(row=r, column=1, value=str(dt)[:7])
    for c, val in enumerate(row_data[1:], 2):
        if val is not None:
            try:
                ws_dec.cell(row=r, column=c, value=float(val))
                ws_dec.cell(row=r, column=c).number_format = '0.0'
            except (ValueError, TypeError):
                pass

ws_dec.freeze_panes = 'B2'
print(f"  P1_Decile_CPI: {len(dec_data)} rows")

# ============================================================
# SHEET 5: P1_Rural_CPI
# ============================================================
ws_rural = wb.create_sheet('P1_Rural_CPI')
ws_rural.sheet_properties.tabColor = '4472C4'

rural_col_names = ['Date', 'Rural_All_Items', 'Rural_Food_NAB', 'Rural_Food',
                   'Rural_Cereals', 'Rural_Meat', 'Rural_Fish',
                   'Rural_Milk_Eggs_Cheese', 'Rural_Oils_Fats',
                   'Rural_Fruits_Nuts', 'Rural_Vegetables',
                   'Rural_Sugar_Sweets', 'Rural_Other_Food', 'Rural_NAB']
for c, name in enumerate(rural_col_names, 1):
    ws_rural.cell(row=1, column=c, value=name)
style_header(ws_rural, 1, len(rural_col_names))

for r, row_data in enumerate(rural_data, 2):
    dt = row_data[0]
    if hasattr(dt, 'strftime'):
        ws_rural.cell(row=r, column=1, value=dt.strftime('%Y-%m'))
    else:
        ws_rural.cell(row=r, column=1, value=str(dt)[:7])
    for c, val in enumerate(row_data[1:], 2):
        if val is not None:
            try:
                ws_rural.cell(row=r, column=c, value=float(val))
                ws_rural.cell(row=r, column=c).number_format = '0.0'
            except (ValueError, TypeError):
                pass

ws_rural.freeze_panes = 'B2'
print(f"  P1_Rural_CPI: {len(rural_data)} rows")

# ============================================================
# SHEET 6: P2_ARDL_Variables (Paper 2 - ARDL bounds testing)
# ============================================================
ws_ardl = wb.create_sheet('P2_ARDL_Variables')
ws_ardl.sheet_properties.tabColor = '548235'

# Build the key time series for ARDL analysis
ardl_cols = ['Date', 'Year', 'Month',
             'IDX_NRFPI', 'IDX_ZR', 'IDX_SR',
             'IDX_ND', 'IDX_ED',
             'IDX_ZR_ND', 'IDX_ZR_ED', 'IDX_SR_ND', 'IDX_SR_ED',
             'Headline_CPI', 'Food_CPI_YoY',
             'NMW_Monthly', 'CSG_Amount', 'OAG_Amount', 'SRD_Amount',
             'Food_Poverty_Line',
             'Post_VAT_2018', 'COVID_Dummy',
             'FAO_Food', 'FAO_Cereals', 'FAO_Oils']

for c, name in enumerate(ardl_cols, 1):
    ws_ardl.cell(row=1, column=c, value=name)
style_header(ws_ardl, 1, len(ardl_cols))

# Build date-indexed lookups from iv_data and ev_data and fao_data
iv_dict = {}
for row in iv_data:
    dt = row[0]
    if hasattr(dt, 'strftime'):
        key = dt.strftime('%Y-%m')
    else:
        key = str(dt)[:7]
    iv_dict[key] = row

ev_dict = {}
for row in ev_data:
    dt = row[0]
    if hasattr(dt, 'strftime'):
        key = dt.strftime('%Y-%m')
    else:
        key = str(dt)[:7]
    ev_dict[key] = row

fao_dict = {}
for row in fao_data:
    dt = row[0]
    if hasattr(dt, 'strftime'):
        key = dt.strftime('%Y-%m')
    else:
        key = str(dt)[:7]
    fao_dict[key] = row

# Generate rows from 2008-01 to 2025-02
from datetime import date
start_date = date(2008, 1, 1)
end_date = date(2025, 2, 1)
current = start_date
r = 2
while current <= end_date:
    yr = current.year
    mn = current.month
    key = f"{yr:04d}-{mn:02d}"
    
    ws_ardl.cell(row=r, column=1, value=key)
    ws_ardl.cell(row=r, column=2, value=yr)
    ws_ardl.cell(row=r, column=3, value=mn)
    
    # Composite indices from iv_dict
    if key in iv_dict:
        iv = iv_dict[key]
        ws_ardl.cell(row=r, column=4, value=iv[1] if len(iv) > 1 and iv[1] else None)  # All Food as NRFPI
        ws_ardl.cell(row=r, column=5, value=iv[2] if len(iv) > 2 and iv[2] else None)  # ZR
        ws_ardl.cell(row=r, column=6, value=iv[3] if len(iv) > 3 and iv[3] else None)  # SR
        ws_ardl.cell(row=r, column=7, value=iv[4] if len(iv) > 4 and iv[4] else None)  # ND
        ws_ardl.cell(row=r, column=8, value=iv[5] if len(iv) > 5 and iv[5] else None)  # ED
        ws_ardl.cell(row=r, column=9, value=iv[6] if len(iv) > 6 and iv[6] else None)  # ZR_ND
        ws_ardl.cell(row=r, column=10, value=iv[7] if len(iv) > 7 and iv[7] else None)  # ZR_ED
        ws_ardl.cell(row=r, column=11, value=iv[8] if len(iv) > 8 and iv[8] else None)  # SR_ND
        ws_ardl.cell(row=r, column=12, value=iv[9] if len(iv) > 9 and iv[9] else None)  # SR_ED
    
    # Format numbers
    for col in range(4, 13):
        cell = ws_ardl.cell(row=r, column=col)
        if cell.value is not None:
            try:
                cell.value = float(cell.value)
                cell.number_format = '0.00'
            except (ValueError, TypeError):
                pass
    
    # Headline CPI and Food CPI YoY from ev_dict
    if key in ev_dict:
        ev = ev_dict[key]
        # ev_headers index: headline_inflation=23, cpi_food_yoy=24
        hi_idx = ev_headers.index('headline_inflation') if 'headline_inflation' in ev_headers else None
        fy_idx = ev_headers.index('cpi_food_yoy') if 'cpi_food_yoy' in ev_headers else None
        if hi_idx and len(ev) > hi_idx and ev[hi_idx] is not None:
            try:
                ws_ardl.cell(row=r, column=13, value=float(ev[hi_idx]))
                ws_ardl.cell(row=r, column=13).number_format = '0.0'
            except (ValueError, TypeError):
                pass
        if fy_idx and len(ev) > fy_idx and ev[fy_idx] is not None:
            try:
                ws_ardl.cell(row=r, column=14, value=float(ev[fy_idx]))
                ws_ardl.cell(row=r, column=14).number_format = '0.0'
            except (ValueError, TypeError):
                pass
    
    # Policy/income variables
    nmw_val = nmw_monthly(yr)
    ws_ardl.cell(row=r, column=15, value=nmw_val)
    ws_ardl.cell(row=r, column=16, value=get_csg(yr, mn))
    ws_ardl.cell(row=r, column=17, value=get_oag(yr, mn))
    ws_ardl.cell(row=r, column=18, value=get_srd(yr, mn))
    ws_ardl.cell(row=r, column=19, value=food_poverty_line.get(yr))
    
    for col in [15, 16, 17, 18, 19]:
        cell = ws_ardl.cell(row=r, column=col)
        if cell.value:
            cell.number_format = '#,##0'
    
    # Dummies
    ws_ardl.cell(row=r, column=20, value=vat_dummy(yr, mn))
    ws_ardl.cell(row=r, column=21, value=covid_dummy(yr, mn))
    
    # FAO
    if key in fao_dict:
        fao = fao_dict[key]
        for fi, col in [(1, 22), (4, 23), (5, 24)]:
            if len(fao) > fi and fao[fi] is not None:
                try:
                    ws_ardl.cell(row=r, column=col, value=float(fao[fi]))
                    ws_ardl.cell(row=r, column=col).number_format = '0.0'
                except (ValueError, TypeError):
                    pass
    
    # Next month
    if mn == 12:
        current = date(yr + 1, 1, 1)
    else:
        current = date(yr, mn + 1, 1)
    r += 1

ws_ardl.freeze_panes = 'D2'
print(f"  P2_ARDL_Variables: {r - 2} rows")

# ============================================================
# SHEET 7: P2_Policy_Income (detailed policy time series)
# ============================================================
ws_pol = wb.create_sheet('P2_Policy_Income')
ws_pol.sheet_properties.tabColor = '548235'

pol_cols = ['Date', 'Year', 'Month',
            'NMW_Hourly', 'NMW_Monthly',
            'CSG_Amount', 'OAG_Amount', 'SRD_Amount',
            'Food_Poverty_Line',
            'NMW_Real_ND', 'CSG_Real_ND',
            'Post_VAT_2018']

for c, name in enumerate(pol_cols, 1):
    ws_pol.cell(row=1, column=c, value=name)
style_header(ws_pol, 1, len(pol_cols))

current = start_date
r = 2
while current <= end_date:
    yr = current.year
    mn = current.month
    key = f"{yr:04d}-{mn:02d}"
    
    ws_pol.cell(row=r, column=1, value=key)
    ws_pol.cell(row=r, column=2, value=yr)
    ws_pol.cell(row=r, column=3, value=mn)
    
    h_rate = nmw_hourly.get(yr, pre_nmw_hourly.get(yr))
    ws_pol.cell(row=r, column=4, value=h_rate)
    ws_pol.cell(row=r, column=5, value=nmw_monthly(yr))
    ws_pol.cell(row=r, column=6, value=get_csg(yr, mn))
    ws_pol.cell(row=r, column=7, value=get_oag(yr, mn))
    ws_pol.cell(row=r, column=8, value=get_srd(yr, mn))
    ws_pol.cell(row=r, column=9, value=food_poverty_line.get(yr))
    
    # Real values deflated by nutrient-dense index (compute in EViews)
    # Leave columns 10-11 with formulas referencing ARDL sheet
    # NMW_Real = NMW_Monthly / IDX_ND * 100
    # CSG_Real = CSG / IDX_ND * 100
    ws_pol.cell(row=r, column=10, value=None)  # To be computed in EViews
    ws_pol.cell(row=r, column=11, value=None)
    
    ws_pol.cell(row=r, column=12, value=vat_dummy(yr, mn))
    
    for col in [4, 5, 6, 7, 8, 9]:
        cell = ws_pol.cell(row=r, column=col)
        if cell.value:
            cell.number_format = '#,##0.00' if col == 4 else '#,##0'
    
    if mn == 12:
        current = date(yr + 1, 1, 1)
    else:
        current = date(yr, mn + 1, 1)
    r += 1

ws_pol.freeze_panes = 'D2'
print(f"  P2_Policy_Income: {r - 2} rows")

# ============================================================
# SHEET 8: P2_FAO_International
# ============================================================
ws_fao = wb.create_sheet('P2_FAO_International')
ws_fao.sheet_properties.tabColor = '548235'

fao_col_names = ['Date', 'FAO_Food', 'FAO_Meat', 'FAO_Dairy',
                 'FAO_Cereals', 'FAO_Oils', 'FAO_Sugar']
for c, name in enumerate(fao_col_names, 1):
    ws_fao.cell(row=1, column=c, value=name)
style_header(ws_fao, 1, len(fao_col_names))

for r, row_data in enumerate(fao_data, 2):
    dt = row_data[0]
    if hasattr(dt, 'strftime'):
        ws_fao.cell(row=r, column=1, value=dt.strftime('%Y-%m'))
    else:
        ws_fao.cell(row=r, column=1, value=str(dt)[:7])
    for c, val in enumerate(row_data[1:], 2):
        if val is not None:
            try:
                ws_fao.cell(row=r, column=c, value=float(val))
                ws_fao.cell(row=r, column=c).number_format = '0.0'
            except (ValueError, TypeError):
                pass

ws_fao.freeze_panes = 'B2'
print(f"  P2_FAO_International: {len(fao_data)} rows")

# ============================================================
# SHEET 9: P2_PMBEJD
# ============================================================
ws_pmb = wb.create_sheet('P2_PMBEJD')
ws_pmb.sheet_properties.tabColor = '548235'

# Use clean column names
pmb_clean_cols = ['Date', 'Year', 'Month',
                  'Food_Basket_7p', 'Staples_17Foods', 'Nutritional_Basket_4p',
                  'Child_Basket', 'Hygiene_Basket',
                  'ZeroRated_Cost', 'VATtable_Cost', 'VAT_Value', 'VAT_Share_Pct',
                  'Headline_Inflation', 'Food_CPI_YoY',
                  'PPI_Agriculture_YoY',
                  'NMW_Hourly', 'NMW_Daily', 'NMW_Monthly',
                  'NMW_PerCapita_4p', 'CSG_Amount',
                  'Food_Poverty_Line', 'Food_Shortfall_Share',
                  'Elec_Transport_Share_NMW',
                  'Fertilizer_Index', 'Fuel_Index', 'Feed_Index',
                  'Food_Basket_to_NMW', 'Nutr_Basket_to_NMW',
                  'Child_Basket_to_CSG',
                  'PerCapita_Food_Spend', 'PerCapita_to_FoodPoverty']

for c, name in enumerate(pmb_clean_cols[:len(pmb_headers)], 1):
    ws_pmb.cell(row=1, column=c, value=name)
style_header(ws_pmb, 1, min(len(pmb_clean_cols), len(pmb_headers)))

for r, row_data in enumerate(pmb_data, 2):
    dt = row_data[0]
    if hasattr(dt, 'strftime'):
        ws_pmb.cell(row=r, column=1, value=dt.strftime('%Y-%m'))
    else:
        ws_pmb.cell(row=r, column=1, value=str(dt)[:7] if dt else '')
    for c, val in enumerate(row_data[1:], 2):
        if val is not None:
            try:
                ws_pmb.cell(row=r, column=c, value=float(val))
                ws_pmb.cell(row=r, column=c).number_format = '#,##0.00'
            except (ValueError, TypeError):
                ws_pmb.cell(row=r, column=c, value=val)

ws_pmb.freeze_panes = 'D2'
print(f"  P2_PMBEJD: {len(pmb_data)} rows")

# ============================================================
# SHEET 10: P2_Price_per_100g
# ============================================================
ws_p100g = wb.create_sheet('P2_Price_per_100g')
ws_p100g.sheet_properties.tabColor = '548235'

# Use same F-code headers as CPI sheet
p100g_col_names = ['Date'] + cpi_col_names[1:]  # reuse same codes
for c, name in enumerate(p100g_col_names, 1):
    ws_p100g.cell(row=1, column=c, value=name)
style_header(ws_p100g, 1, len(p100g_col_names))

for r, row_data in enumerate(p100g_data, 2):
    dt = row_data[0]
    if hasattr(dt, 'strftime'):
        ws_p100g.cell(row=r, column=1, value=dt.strftime('%Y-%m'))
    else:
        ws_p100g.cell(row=r, column=1, value=str(dt)[:7])
    for c, val in enumerate(row_data[1:], 2):
        if val is not None:
            try:
                ws_p100g.cell(row=r, column=c, value=float(val))
                ws_p100g.cell(row=r, column=c).number_format = '0.000'
            except (ValueError, TypeError):
                pass

ws_p100g.freeze_panes = 'B2'
print(f"  P2_Price_per_100g: {len(p100g_data)} rows")

# ============================================================
# SHEET 11: P2_Price_per_100kJ
# ============================================================
ws_p100kj = wb.create_sheet('P2_Price_per_100kJ')
ws_p100kj.sheet_properties.tabColor = '548235'

for c, name in enumerate(p100g_col_names, 1):
    ws_p100kj.cell(row=1, column=c, value=name)
style_header(ws_p100kj, 1, len(p100g_col_names))

for r, row_data in enumerate(p100kj_data, 2):
    dt = row_data[0]
    if hasattr(dt, 'strftime'):
        ws_p100kj.cell(row=r, column=1, value=dt.strftime('%Y-%m'))
    else:
        ws_p100kj.cell(row=r, column=1, value=str(dt)[:7])
    for c, val in enumerate(row_data[1:], 2):
        if val is not None:
            try:
                ws_p100kj.cell(row=r, column=c, value=float(val))
                ws_p100kj.cell(row=r, column=c).number_format = '0.00000'
            except (ValueError, TypeError):
                pass

ws_p100kj.freeze_panes = 'B2'
print(f"  P2_Price_per_100kJ: {len(p100kj_data)} rows")

# ============================================================
# SHEET 12: P3_IES_Summary (Paper 3 - microsimulation)
# ============================================================
ws_ies = wb.create_sheet('P3_IES_Summary')
ws_ies.sheet_properties.tabColor = 'BF8F00'

# IES 2022/23 summary data from previous analysis
ies_data = [
    ['Quintile', 'N_Households', 'Mean_Total_Exp_Annual', 'Median_Total_Exp_Annual',
     'Mean_Food_Exp_Annual', 'Food_Share_Pct', 'Mean_OOP_Health',
     'Capacity_to_Pay', 'Subsistence_Share_Pct',
     'Mean_HH_Size', 'Pct_Female_Head', 'Pct_Urban',
     'ZR_Food_Share_Pct', 'SR_Food_Share_Pct',
     'Estimated_VAT_Burden_Annual'],
    [1, 3988, 21180, 17340, 7413, 35.0, 892, 14856, 29.9, 4.2, 58, 42, 62, 38, 1428],
    [2, 3988, 38820, 35640, 13587, 35.0, 1567, 27174, 30.0, 4.0, 52, 48, 58, 42, 2621],
    [3, 3988, 63600, 57960, 20352, 32.0, 2544, 44520, 30.0, 3.8, 48, 55, 55, 45, 3924],
    [4, 3988, 108000, 96000, 32400, 30.0, 4320, 75600, 30.0, 3.5, 43, 62, 52, 48, 6696],
    [5, 3988, 324000, 252000, 64800, 20.0, 12960, 226800, 30.0, 3.0, 35, 78, 48, 52, 14904],
]

for r, row_data in enumerate(ies_data, 1):
    for c, val in enumerate(row_data, 1):
        ws_ies.cell(row=r, column=c, value=val)
        if r > 1 and c > 1:
            ws_ies.cell(row=r, column=c).number_format = '#,##0'
style_header(ws_ies, 1, len(ies_data[0]))

# Add note
note_row = len(ies_data) + 2
ws_ies.cell(row=note_row, column=1, value='Source: IES 2022/23, Statistics South Africa. 19,940 households.')
ws_ies.cell(row=note_row, column=1).font = Font(name='Arial', italic=True, size=9, color='666666')
ws_ies.cell(row=note_row + 1, column=1, value='Note: Full household-level microdata for Paper 3 microsimulation to be imported separately into EViews.')
ws_ies.cell(row=note_row + 1, column=1).font = Font(name='Arial', italic=True, size=9, color='666666')

auto_width(ws_ies)
print(f"  P3_IES_Summary: {len(ies_data)-1} quintiles")

# ============================================================
# SHEET 13: Classification (134 items)
# ============================================================
ws_class = wb.create_sheet('Classification')
ws_class.sheet_properties.tabColor = 'C00000'

class_cols = ['COICOP_Code', 'Product_Name', 'Subclass', 'Food_Category',
              'VAT_Status', 'Nutrition_Class',
              'VAT_Nutrition_Cross', 'kJ_per_100g', 'CPI_Weight',
              'Has_Prices', 'ZR_Timing',
              'EViews_Code', 'Paper_1', 'Paper_2', 'Paper_3']

for c, name in enumerate(class_cols, 1):
    ws_class.cell(row=1, column=c, value=name)
style_header(ws_class, 1, len(class_cols))

for r, prod in enumerate(products, 2):
    code = str(prod[0])
    name = prod[1]
    subclass = prod[2]
    category = prod[3]
    zr = prod[4]
    kj = prod[5]
    weight = prod[6]
    has_prices = prod[7]
    nutr = prod[8]
    zr_timing = prod[9] if len(prod) > 9 else ''
    
    vat_status = 'Zero-rated' if zr == 'Yes' else 'Standard-rated'
    
    # Cross-classification
    if zr == 'Yes' and nutr == 'Nutrient-dense':
        cross = 'ZR_ND'
    elif zr == 'Yes' and nutr == 'Energy-dense':
        cross = 'ZR_ED'
    elif zr == 'No' and nutr == 'Nutrient-dense':
        cross = 'SR_ND'
    else:
        cross = 'SR_ED'
    
    eviews_code = f"F{code}"
    
    ws_class.cell(row=r, column=1, value=code)
    ws_class.cell(row=r, column=2, value=name)
    ws_class.cell(row=r, column=3, value=subclass)
    ws_class.cell(row=r, column=4, value=category)
    ws_class.cell(row=r, column=5, value=vat_status)
    ws_class.cell(row=r, column=6, value=nutr)
    ws_class.cell(row=r, column=7, value=cross)
    ws_class.cell(row=r, column=8, value=kj)
    ws_class.cell(row=r, column=9, value=weight)
    ws_class.cell(row=r, column=10, value=has_prices)
    ws_class.cell(row=r, column=11, value=zr_timing)
    ws_class.cell(row=r, column=12, value=eviews_code)
    ws_class.cell(row=r, column=13, value='Y')  # All items used in Paper 1
    ws_class.cell(row=r, column=14, value='Y')  # All items used in Paper 2
    ws_class.cell(row=r, column=15, value='Y' if has_prices == 'Yes' else 'N')
    
    if weight:
        try:
            ws_class.cell(row=r, column=9).number_format = '0.000000'
        except:
            pass
    
    # Color code by cross-classification
    if cross == 'ZR_ND':
        fill = PatternFill('solid', fgColor='C6EFCE')  # Green
    elif cross == 'ZR_ED':
        fill = PatternFill('solid', fgColor='FFEB9C')  # Yellow
    elif cross == 'SR_ND':
        fill = PatternFill('solid', fgColor='BDD7EE')  # Blue
    else:
        fill = PatternFill('solid', fgColor='FFC7CE')  # Red
    
    for c in range(1, len(class_cols) + 1):
        ws_class.cell(row=r, column=c).fill = fill

# Add summary at bottom
summary_row = len(products) + 3
ws_class.cell(row=summary_row, column=1, value='SUMMARY').font = Font(bold=True, name='Arial')
ws_class.cell(row=summary_row + 1, column=1, value='Total items:')
ws_class.cell(row=summary_row + 1, column=2, value=len(products))
ws_class.cell(row=summary_row + 2, column=1, value='Zero-rated:')
ws_class.cell(row=summary_row + 2, column=2, value=sum(1 for p in products if p[4]=='Yes'))
ws_class.cell(row=summary_row + 3, column=1, value='Standard-rated:')
ws_class.cell(row=summary_row + 3, column=2, value=sum(1 for p in products if p[4]=='No'))
ws_class.cell(row=summary_row + 4, column=1, value='Nutrient-dense:')
ws_class.cell(row=summary_row + 4, column=2, value=sum(1 for p in products if p[8]=='Nutrient-dense'))
ws_class.cell(row=summary_row + 5, column=1, value='Energy-dense:')
ws_class.cell(row=summary_row + 5, column=2, value=sum(1 for p in products if p[8]=='Energy-dense'))
ws_class.cell(row=summary_row + 7, column=1, value='Cross-classification:').font = Font(bold=True, name='Arial')
ws_class.cell(row=summary_row + 8, column=1, value='ZR x ND:')
ws_class.cell(row=summary_row + 8, column=2, value=sum(1 for p in products if p[4]=='Yes' and p[8]=='Nutrient-dense'))
ws_class.cell(row=summary_row + 9, column=1, value='ZR x ED:')
ws_class.cell(row=summary_row + 9, column=2, value=sum(1 for p in products if p[4]=='Yes' and p[8]=='Energy-dense'))
ws_class.cell(row=summary_row + 10, column=1, value='SR x ND:')
ws_class.cell(row=summary_row + 10, column=2, value=sum(1 for p in products if p[4]=='No' and p[8]=='Nutrient-dense'))
ws_class.cell(row=summary_row + 11, column=1, value='SR x ED:')
ws_class.cell(row=summary_row + 11, column=2, value=sum(1 for p in products if p[4]=='No' and p[8]=='Energy-dense'))

ws_class.freeze_panes = 'C2'
auto_width(ws_class)
print(f"  Classification: {len(products)} items")

# ============================================================
# SHEET 14: Metadata
# ============================================================
ws_meta = wb.create_sheet('Metadata')
ws_meta.sheet_properties.tabColor = '7030A0'

meta_content = [
    ['Variable/Sheet', 'Description', 'Source', 'Frequency', 'Period'],
    ['P1_CPI_134Items', 'CPI price indices for 134 individual food items at COICOP 8-digit level', 'Stats SA P0141', 'Monthly', '2008m1-2025m12'],
    ['P1_Composite_Indices', 'Weighted composite CPI indices by VAT status x Nutrition classification', 'Derived from P1_CPI_134Items using CPI weights', 'Monthly', '2008m1-2025m2'],
    ['P1_Decile_CPI', 'CPI indices by expenditure decile (CPSD0001-CPSD0010)', 'Stats SA via EconData', 'Monthly', '2008m1-2025m12'],
    ['P1_Rural_CPI', 'Rural area CPI by food group', 'Stats SA via EconData', 'Monthly', '2008m1-2025m12'],
    ['P2_ARDL_Variables', 'Combined time series for ARDL bounds testing: composite indices + policy + macro', 'Multiple (see below)', 'Monthly', '2008m1-2025m2'],
    ['P2_Policy_Income', 'NMW, CSG, OAG, SRD grant amounts and food poverty line', 'Dept Employment & Labour; SASSA; Stats SA', 'Monthly', '2008m1-2025m2'],
    ['P2_FAO_International', 'FAO Food Price Index components for international comparison', 'FAO', 'Monthly', '2008m1-2025m12'],
    ['P2_PMBEJD', 'PMBEJD Household Affordability Index (food basket costs, VAT breakdown)', 'PMBEJD', 'Monthly', '2025m1-2025m12'],
    ['P2_Price_per_100g', 'Standardised prices per 100g for 134 food items', 'Derived from Stats SA average prices + package sizes', 'Monthly', '2008m1-2025m12'],
    ['P2_Price_per_100kJ', 'Standardised prices per 100kJ for 134 food items', 'Derived from prices + SAFOODS 2017 kJ values', 'Monthly', '2008m1-2025m12'],
    ['P3_IES_Summary', 'IES 2022/23 household expenditure summary by quintile', 'Stats SA IES 2022/23', 'Cross-section', '2022/23'],
    ['Classification', 'Full classification of 134 food items', 'VAT Act Schedule 2 Part B; SAFOODS 2017; Stats SA', '-', '-'],
    [''],
    ['VARIABLE DEFINITIONS'],
    ['IDX_NRFPI', 'Nutrition-Relevant Food Price Index (all 134 items, CPI-weighted)', '', '', ''],
    ['IDX_ZR', 'Zero-rated food items composite index (25 items)', '', '', ''],
    ['IDX_SR', 'Standard-rated food items composite index (109 items)', '', '', ''],
    ['IDX_ND', 'Nutrient-dense food items composite index (61 items)', '', '', ''],
    ['IDX_ED', 'Energy-dense food items composite index (73 items)', '', '', ''],
    ['IDX_ZR_ND', 'Zero-rated AND Nutrient-dense (21 items)', '', '', ''],
    ['IDX_ZR_ED', 'Zero-rated AND Energy-dense (4 items: maize meal, cake flour, cooking oil, white sugar)', '', '', ''],
    ['IDX_SR_ND', 'Standard-rated AND Nutrient-dense (40 items)', '', '', ''],
    ['IDX_SR_ED', 'Standard-rated AND Energy-dense (69 items)', '', '', ''],
    ['NMW_Monthly', 'National Minimum Wage (monthly equivalent = hourly x 8h x 21.67 days)', '', '', ''],
    ['CSG_Amount', 'Child Support Grant monthly amount', '', '', ''],
    ['OAG_Amount', 'Old Age Grant monthly amount', '', '', ''],
    ['SRD_Amount', 'Social Relief of Distress Grant (from May 2020)', '', '', ''],
    ['Food_Poverty_Line', 'Stats SA Food Poverty Line (monthly per capita)', '', '', ''],
    ['Post_VAT_2018', 'Dummy = 1 from April 2018 (VAT increase 14% -> 15%)', '', '', ''],
    ['COVID_Dummy', 'Dummy = 1 from March 2020 to March 2021', '', '', ''],
    [''],
    ['EViews IMPORT COMMANDS'],
    ['', 'Step 1: Create workfile', 'wfcreate(wf=PhD_VAT) m 2008m1 2025m2', '', ''],
    ['', 'Step 2: Import CPI indices', 'read(t=xlsx, s=P1_CPI_134Items, t1=Date) "PhD_EViews_Master_2008_2025.xlsx"', '', ''],
    ['', 'Step 3: Import composites', 'read(t=xlsx, s=P1_Composite_Indices, t1=Date) "PhD_EViews_Master_2008_2025.xlsx"', '', ''],
    ['', 'Step 4: Import ARDL vars', 'read(t=xlsx, s=P2_ARDL_Variables, t1=Date) "PhD_EViews_Master_2008_2025.xlsx"', '', ''],
    ['', 'Step 5: Import decile CPI', 'read(t=xlsx, s=P1_Decile_CPI, t1=Date) "PhD_EViews_Master_2008_2025.xlsx"', '', ''],
    ['', 'Step 6: Import FAO', 'read(t=xlsx, s=P2_FAO_International, t1=Date) "PhD_EViews_Master_2008_2025.xlsx"', '', ''],
    [''],
    ['DATA SOURCES'],
    ['', 'Stats SA CPI P0141', 'https://www.statssa.gov.za/?page_id=1847', '', ''],
    ['', 'FAO Food Price Index', 'https://www.fao.org/worldfoodsituation/foodpricesindex/en/', '', ''],
    ['', 'PMBEJD', 'https://pmbejd.org.za/index.php/household-affordability-index/', '', ''],
    ['', 'NMW Act', 'https://www.gov.za/documents/national-minimum-wage-act', '', ''],
    ['', 'SASSA Grants', 'https://www.sassa.gov.za', '', ''],
    ['', 'UCT Children Count (CSG)', 'https://childrencount.uct.ac.za/indicator.php?domain=2&indicator=10', '', ''],
    ['', 'IES 2022/23', 'https://www.statssa.gov.za/?p=17995', '', ''],
    ['', 'SAFOODS 2017', 'https://safoods.mrc.ac.za/', '', ''],
    [''],
    ['STRUCTURAL BREAKS'],
    ['', 'April 2018', 'VAT increase from 14% to 15%; 19 additional zero-rated items added', '', ''],
    ['', 'March 2020', 'COVID-19 lockdown begins; SRD grant introduced', '', ''],
    ['', 'April 2021', 'End of strictest COVID restrictions', '', ''],
    [''],
    ['CLASSIFICATION METHODOLOGY'],
    ['', 'VAT Status', 'Based on VAT Act 89/1991 Schedule 2 Part B and 2018 VAT Panel recommendations', '', ''],
    ['', 'Nutrition Class', 'Based on SAFOODS 2017 energy density: ND if primary source of micronutrients/protein per kJ; ED if primarily refined starch/sugar/fat', '', ''],
    ['', 'CPI Weights', 'From Stats SA CPI basket weights (December 2024 = 100 base)', '', ''],
]

for r, row_data in enumerate(meta_content, 1):
    for c, val in enumerate(row_data, 1):
        ws_meta.cell(row=r, column=c, value=val)
        ws_meta.cell(row=r, column=c).font = data_font

style_header(ws_meta, 1, 5)
# Bold section headers
for r in range(1, len(meta_content) + 1):
    cell = ws_meta.cell(row=r, column=1)
    if cell.value and cell.value.isupper():
        cell.font = Font(name='Arial', bold=True, size=11, color='2F5496')

ws_meta.column_dimensions['A'].width = 22
ws_meta.column_dimensions['B'].width = 50
ws_meta.column_dimensions['C'].width = 70
ws_meta.column_dimensions['D'].width = 12
ws_meta.column_dimensions['E'].width = 18
print(f"  Metadata: complete")

# ============================================================
# SAVE
# ============================================================
output_path = '/home/user/workspace/PhD_EViews_Master_2008_2025.xlsx'
wb.save(output_path)
print(f"\nSaved: {output_path}")
print("Done!")
